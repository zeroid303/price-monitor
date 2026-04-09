"""
3사 크롤링 데이터 → 비교 엑셀 생성
기존 경쟁사_가격비교_전수조사 양식 기반
명함천국 전 제품을 기준으로 비즈하우스/오프린트미 매칭
"""
import json
import glob
import re
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))


def load_latest(prefix):
    files = sorted(glob.glob(f'output/{prefix}_*.json'))
    if not files:
        return []
    with open(files[-1], 'r', encoding='utf-8') as f:
        return [r for r in json.load(f) if not r.get('error')]


def interpolate_200(p100, p250):
    if p100 and p250:
        return round(p100 + (p250 - p100) * 100 / 150)
    return None


# ── 매칭 테이블 ──
# 명함천국 용지 → (비즈하우스 상품, 비즈하우스 용지, 오프린트미 key, 비고)
# 비즈하우스는 "상품명|용지명" 형식으로 검색
BASIC_MATCH = {
    '스노우화이트 216g / 코팅': (None, None, None, ''),
    '스노우화이트 216g / 코팅 [일반코팅명함]': (None, None, None, ''),
    '스노우화이트 250g / 비코팅': ('기본명함', '스노우 250g', 'bestBusinessCard/soft', ''),
    '스노우화이트 250g / 무광코팅': ('기본명함', '코팅스노우 219g', None, ''),
    '스노우화이트 300g / 무광코팅': ('프리미엄', '무코팅스노우 300g', None, ''),
    '스노우화이트 300g / 유광코팅': (None, None, None, ''),
    '스노우화이트 400g / 무광코팅': ('프리미엄', '무코팅스노우 400g', None, ''),
    '누브지 209g / 비코팅': (None, None, None, '명함천국 전용'),
    '휘라레 216g / 비코팅': ('재질', '명함 휘라레 216g', 'premiumBusinessCard/linen', ''),
    '반누보화이트 250g / 비코팅': ('기본명함', '반누보 210g', 'bestBusinessCard/premium_matte', '평량 다름(250g vs 210g)'),
    '머쉬멜로우 209g / 비코팅': (None, None, None, '명함천국 전용'),
    '스코틀랜드 220g / 비코팅': (None, None, 'premiumBusinessCard/felt', '오프린트미: 펠트(유사)'),
    '유포지 250g / 비코팅': ('기본명함', '유포지 250g', None, ''),
    '스타드림 240g / 비코팅': ('펄지', '명함 스타드림골드 240g', 'premiumBusinessCard/pearl', ''),
    '키칼라메탈릭 200g / 비코팅': (None, None, None, ''),
    '크리스탈펄(팝셋대체) 235g / 비코팅': ('펄지', '명함 팝셋화이트 240g', None, '유사 펄지'),
    '컨셉 285g / 비코팅': ('펄지', '명함 컨셉 블루펄 285g', None, ''),
    '스타골드 250g / 비코팅': (None, None, None, ''),
    '카멜레온 200g / 비코팅': (None, None, None, ''),
    '키칼라골드 200g / 비코팅': (None, None, None, ''),
    '그레이스 256g / 비코팅': (None, None, None, ''),
    '크라프트팩 250g / 비코팅': (None, None, 'ecoBusinessCard/kraft', ''),
    '아르떼 310g / 비코팅': (None, None, None, ''),
    '그문드 바우하우스 250g / 비코팅': (None, None, None, ''),
    '엑스트라 누브 350g / 비코팅': ('프리미엄', '명함 엑스트라 누보 350g', None, ''),
    '엑스트라 머쉬 350g / 비코팅': ('프리미엄', '명함 엑스트라 매쉬멜로우 350g', None, ''),
    '엑스트라 린넨 350g / 비코팅': (None, None, None, ''),
    '엑스트라 에그화이트 400g / 비코팅': (None, None, None, ''),
    '엑스트라 스타드림 340g / 비코팅': (None, None, None, ''),
    '엑스트라 띤또레또 350g / 비코팅': (None, None, 'premiumBusinessCard/felt', '오프린트미: 펠트(유사)'),
    '엑스트라 매트화이트 350g / 비코팅': (None, None, None, ''),
}

# 특수명함 매칭: (명함천국 상품명, 비즈하우스 상품명 키워드, 오프린트미 key, 비고)
SPECIAL_MATCH = {
    '2단명함': ('2단', None, ''),
    '카드명함': ('카드 명함', None, ''),
    '에폭시명함': (None, None, '경쟁사 없음'),
    '부분코팅명함': (None, None, '경쟁사 없음'),
    '화이트명함': (None, 'premiumBusinessCard/matte_black', '오프린트미: 매트블랙(유사)'),
    '형광명함': (None, None, '경쟁사 없음'),
    '엣지명함': (None, None, '경쟁사 없음'),
    'PET카드명함': (None, 'transPremiumBusinessCard/transparent', '오프린트미: 투명명함'),
    '3D금박명함': ('3D명함', 'premiumBusinessCard/gold_glossy', '스코딕스/UV엠보싱'),
}


def build_excel():
    ec_data = load_latest('Ecard21Crawler')
    bz_data = load_latest('BizhowsFullCrawler')
    oh_data = load_latest('OhprintCrawler')

    print(f"명함천국: {len(ec_data)}건, 비즈하우스: {len(bz_data)}건, 오프린트미: {len(oh_data)}건")

    # ── 비즈하우스 인덱스: "상품키워드|용지" → price ──
    bz_index = {}
    for r in bz_data:
        if r.get('category') != '명함':
            continue
        bz_index[r['spec']] = r['price']

    def find_bz_price(product_keyword, paper_keyword):
        """비즈하우스 가격 찾기 (spec에서 부분 매칭)"""
        if not product_keyword:
            return None, ''
        for spec, price in bz_index.items():
            if product_keyword in spec and (not paper_keyword or paper_keyword in spec):
                return price, spec.replace(' 200매', '')
        return None, ''

    # ── 오프린트미 인덱스: key → 200매 보간 가격 ──
    oh_by_key = {}
    for r in oh_data:
        km = re.search(r'\[key=([^\]]+)\]', r['spec'])
        qm = re.search(r'(\d+)매', r['spec'])
        nm = re.match(r'(.+?)\s+\d+매', r['spec'])
        if km and qm:
            key = km.group(1)
            qty = int(qm.group(1))
            name = nm.group(1).replace(f' [key={key}]', '') if nm else key
            oh_by_key.setdefault(key, {'name': name, 'qtys': {}})['qtys'][qty] = r['price']

    oh_products = {}
    for key, info in oh_by_key.items():
        p200 = interpolate_200(info['qtys'].get(100), info['qtys'].get(250))
        oh_products[key] = {'name': info['name'], 'price_200': p200}

    def find_oh_price(oh_key):
        if not oh_key or oh_key not in oh_products:
            return None, ''
        info = oh_products[oh_key]
        return info['price_200'], info['name']

    # ── 명함천국 데이터 정리 ──
    ec_items = []
    for r in ec_data:
        if r.get('category') != '명함':
            continue
        spec = r['spec']
        # 순위 표시 항목 제외
        if re.search(r'\d+위\.', spec):
            continue
        m = re.match(r'(\S+명함)\s+(.+?)\s+200매', spec)
        if m:
            ec_items.append({
                'type': m.group(1),
                'paper': m.group(2),
                'price': r['price'],
                'spec': spec,
            })

    # ── 엑셀 생성 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "명함 비교"

    # 스타일
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(name='Arial', size=10, bold=True)
    cell_font = Font(name='Arial', size=10)
    section_font = Font(name='Arial', size=10, bold=True)
    section_fill = PatternFill('solid', fgColor='F2F2F2')
    blue_fill = PatternFill('solid', fgColor='DBEEF4')    # 자사
    green_fill = PatternFill('solid', fgColor='E2EFDA')   # 경쟁사 저렴
    red_fill = PatternFill('solid', fgColor='FCE4EC')     # 경쟁사 비쌈
    price_fmt = '#,##0'

    headers = [
        '구분', '명함천국 상품/용지', '수량',
        '명함천국\n(VAT포함)', '명함천국\n(VAT제외)',
        '오프린트미\n(VAT별도)', '오프린트미 상품',
        '비즈하우스\n(VAT별도)', '비즈하우스 상품',
        '비고',
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 2

    def write_section(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = section_font
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = section_fill
        row += 1

    def write_row(ptype, paper, ec_price, oh_price, oh_name, bz_price, bz_name, note):
        nonlocal row
        ws.cell(row=row, column=1, value=ptype).font = cell_font
        ws.cell(row=row, column=2, value=paper).font = cell_font
        ws.cell(row=row, column=3, value='200매').font = cell_font

        # 명함천국 가격
        c4 = ws.cell(row=row, column=4, value=ec_price)
        c4.font = cell_font
        c4.number_format = price_fmt
        c4.fill = blue_fill
        vat_ex = round(ec_price / 1.1) if ec_price else None
        c5 = ws.cell(row=row, column=5, value=vat_ex)
        c5.font = cell_font
        c5.number_format = price_fmt
        c5.fill = blue_fill

        # 오프린트미
        if oh_price:
            c6 = ws.cell(row=row, column=6, value=oh_price)
            c6.font = cell_font
            c6.number_format = price_fmt
            # 색상: VAT제외 기준 비교
            if vat_ex and oh_price < vat_ex:
                c6.fill = green_fill
            elif vat_ex and oh_price > vat_ex:
                c6.fill = red_fill
        else:
            ws.cell(row=row, column=6, value='-').font = cell_font

        ws.cell(row=row, column=7, value=oh_name or '').font = cell_font

        # 비즈하우스
        if bz_price:
            c8 = ws.cell(row=row, column=8, value=bz_price)
            c8.font = cell_font
            c8.number_format = price_fmt
            if vat_ex and bz_price < vat_ex:
                c8.fill = green_fill
            elif vat_ex and bz_price > vat_ex:
                c8.fill = red_fill
        else:
            ws.cell(row=row, column=8, value='-').font = cell_font

        ws.cell(row=row, column=9, value=bz_name or '').font = cell_font
        ws.cell(row=row, column=10, value=note or '').font = cell_font
        row += 1

    # ── 기본명함 ──
    write_section('기본명함 (용지별, 200매)')
    basic_items = [i for i in ec_items if i['type'] == '기본명함']
    for item in sorted(basic_items, key=lambda x: x['price']):
        paper = item['paper']
        match = BASIC_MATCH.get(paper, (None, None, None, ''))
        bz_prod, bz_paper, oh_key, note = match

        bz_price, bz_name = find_bz_price(bz_prod, bz_paper)
        oh_price, oh_name = find_oh_price(oh_key)
        if oh_price:
            oh_name = f"{oh_name} 200매(보간)"

        write_row('기본명함', paper, item['price'], oh_price, oh_name, bz_price, bz_name, note)

    # ── 특수명함 ──
    write_section('특수명함 (200매)')
    special_types = ['2단명함', '에폭시명함', '부분코팅명함', '화이트명함',
                     '형광명함', '엣지명함', 'PET카드명함', '3D금박명함', '카드명함']

    for stype in special_types:
        items = [i for i in ec_items if i['type'] == stype]
        if not items:
            continue

        smatch = SPECIAL_MATCH.get(stype, (None, None, ''))
        bz_keyword, oh_key, base_note = smatch

        for item in sorted(items, key=lambda x: x['price']):
            paper = item['paper']

            # 비즈하우스 매칭
            bz_price, bz_name = None, ''
            if bz_keyword:
                bz_price, bz_name = find_bz_price(bz_keyword, None)

            # 오프린트미 매칭
            oh_price, oh_name = find_oh_price(oh_key)
            if oh_price:
                oh_name = f"{oh_name} 200매(보간)"

            note = base_note
            write_row(stype, paper, item['price'], oh_price, oh_name, bz_price, bz_name, note)

    # ── 범례 ──
    row += 1
    write_section('[범례]')
    legends = [
        f'  기준일: {datetime.now().strftime("%Y-%m-%d")}',
        '  파란배경 = 명함천국 자사 가격 (VAT제외 = VAT포함가 / 1.1)',
        '  초록배경 = 경쟁사가 명함천국보다 저렴',
        '  빨간배경 = 경쟁사가 명함천국보다 비쌈',
        '  회색(-) = 해당 사이트에 비교 상품 없음',
        '  명함천국: Playwright 크롤링 (200매 VAT포함)',
        '  오프린트미: Playwright 크롤링 (VAT별도, 200매 보간=100매~250매 선형추정)',
        '  비즈하우스: Playwright 크롤링 (VAT별도, 200매)',
        f'  명함천국 {len([i for i in ec_items if i["type"]=="기본명함"])}개 기본용지 + 특수명함 {len([i for i in ec_items if i["type"]!="기본명함"])}건',
        f'  비즈하우스 {len(bz_data)}건 (카테고리 페이지 전체 상품 × 용지)',
        f'  오프린트미 {len(oh_products)}종 (200매 보간)',
    ]
    for legend in legends:
        ws.cell(row=row, column=1, value=legend).font = cell_font
        row += 1

    # 열 너비
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 35
    ws.column_dimensions['J'].width = 30

    # 저장
    today = datetime.now().strftime("%y%m%d")
    filepath = f"C:/Users/Admin/Downloads/경쟁사_가격비교_전수조사_{today}.xlsx"
    wb.save(filepath)

    basic_count = len([i for i in ec_items if i['type'] == '기본명함'])
    special_count = len([i for i in ec_items if i['type'] != '기본명함'])
    print(f"\n저장 완료: {filepath}")
    print(f"  기본명함: {basic_count}종")
    print(f"  특수명함: {special_count}건")
    print(f"  총 {basic_count + special_count}행")


if __name__ == "__main__":
    build_excel()
