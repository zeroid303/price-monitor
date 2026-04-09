import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load data
with open("C:/Workspace/my-project/analysis/price-monitor/output/ecard21_flyer_prices.json", "r", encoding="utf-8") as f:
    ecard_data = json.load(f)

with open("C:/Workspace/my-project/analysis/price-monitor/output/bizhows_flyer_prices.json", "r", encoding="utf-8") as f:
    biz_data = json.load(f)


def parse_biz_price(price_str):
    price_str = price_str.replace(",", "").replace("\uc6d0", "")
    m = re.findall(r"(\d+)", price_str)
    if len(m) >= 3:
        return int(m[-1])
    elif len(m) >= 1:
        return int(m[0])
    return 0


def normalize_biz_paper(paper_str):
    paper_str = paper_str.strip()
    patterns = [
        r"(\uc544\ud2b8\uc9c0)\s*(\d+)g",
        r"(\uc2a4\ub178\uc6b0\uc9c0)\s*(\d+)g",
        r"(\ubaa8\uc870\uc9c0)\s*(\d+)g",
    ]
    for pat in patterns:
        m = re.search(pat, paper_str)
        if m:
            return m.group(1), int(m.group(2))
    paren = re.search(r"\(([^)]+)\)", paper_str)
    if paren:
        inner = paren.group(1)
        for pat in patterns:
            m = re.search(pat, inner)
            if m:
                return m.group(1), int(m.group(2))
    return None, None


def get_biz_side(item):
    product = item.get("product", "")
    side = item.get("side", "")
    if "\ub2e8\uba74" in product or side == "\ub2e8\uba74":
        return "\ub2e8\uba74"
    elif "\uc591\uba74" in product or side == "\uc591\uba74":
        return "\uc591\uba74"
    return "\uc591\uba74"


def normalize_ecard_paper(paper_name):
    patterns = [
        (r"\uc544\ud2b8\uc9c0\s*(\d+)g", "\uc544\ud2b8\uc9c0"),
        (r"\uc2a4\ub178\uc6b0\ud654\uc774\ud2b8\s*(\d+)g", "\uc2a4\ub178\uc6b0\uc9c0"),
        (r"\ubaa8\uc870\uc9c0\s*(\d+)g", "\ubaa8\uc870\uc9c0"),
    ]
    for pat, ptype in patterns:
        m = re.search(pat, paper_name)
        if m:
            return ptype, int(m.group(1))
    return None, None


ecard_products = ecard_data["products"]

# Build ecard21 index: (product_type, paper_type, gram) -> data (first size variant only)
ecard_index = {}
for product_type, items in ecard_products.items():
    for item in items:
        ptype, gram = normalize_ecard_paper(item["paper_name"])
        if ptype is None:
            continue
        key = (product_type, ptype, gram)
        if key not in ecard_index:
            ecard_index[key] = item

# Build bizhows A4 items
biz_a4_items = []
for item in biz_data:
    size = item.get("size", "")
    if "A4" not in size and "21.0" not in size:
        continue
    paper_type, gram = normalize_biz_paper(item.get("paper", ""))
    if paper_type is None:
        continue
    side = get_biz_side(item)
    biz_a4_items.append({
        "category": item["category"],
        "product": item["product"],
        "paper": item["paper"],
        "paper_type": paper_type,
        "gram": gram,
        "side": side,
        "qty": int(item["qty"]),
        "price": parse_biz_price(item["price"]),
        "price_raw": item["price"],
        "url": item["url"],
    })

# Group ecard by paper
ecard_by_paper = {}
for (product_type, ptype, gram), item in ecard_index.items():
    paper_key = (ptype, gram)
    if paper_key not in ecard_by_paper:
        ecard_by_paper[paper_key] = {}
    ecard_by_paper[paper_key][product_type] = item

# Group bizhows by paper
biz_by_paper = {}
for item in biz_a4_items:
    paper_key = (item["paper_type"], item["gram"])
    if paper_key not in biz_by_paper:
        biz_by_paper[paper_key] = {}
    prod_side_key = (item["product"], item["side"])
    if prod_side_key not in biz_by_paper[paper_key]:
        biz_by_paper[paper_key][prod_side_key] = item

# All unique ecard papers
all_ecard_papers = set()
for (product_type, ptype, gram) in ecard_index.keys():
    all_ecard_papers.add((ptype, gram))

paper_order = {"\uc544\ud2b8\uc9c0": 1, "\uc2a4\ub178\uc6b0\uc9c0": 2, "\ubaa8\uc870\uc9c0": 3}
sorted_papers = sorted(all_ecard_papers, key=lambda x: (paper_order.get(x[0], 99), x[1]))


def ecard_paper_display(ptype, gram):
    if ptype == "\uc2a4\ub178\uc6b0\uc9c0":
        return f"\uc2a4\ub178\uc6b0\ud654\uc774\ud2b8 {gram}g"
    return f"{ptype} {gram}g"


# --- Build reference JSON ---
matched_count = 0
papers_list = []

for ptype, gram in sorted_papers:
    ecard_pname = ecard_paper_display(ptype, gram)
    biz_pname = f"{ptype} {gram}g"

    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    biz_prods = biz_by_paper.get((ptype, gram), {})

    paper_entry = {
        "paper_id": f"{ptype}_{gram}g",
        "paper_name_ko": ecard_pname,
        "ecard21": {"paper_name": ecard_pname, "prices": {}},
        "bizhows": {"paper_name": biz_pname if biz_prods else None, "prices": {}},
    }

    for prod_type, item in ecard_prods.items():
        price_key = f"{prod_type.replace(' ', '_')}_\uc591\uba74_{item['quantity']}"
        paper_entry["ecard21"]["prices"][price_key] = item["price"]
        paper_entry["ecard21"]["paper_code"] = item.get("paper_code", "")

    for (prod_name, side), item in biz_prods.items():
        price_key = f"{prod_name.replace(' ', '_')}_{side}_{item['qty']}"
        paper_entry["bizhows"]["prices"][price_key] = item["price"]

    if biz_prods:
        paper_entry["match_confidence"] = "high"
        matched_count += 1
    else:
        paper_entry["match_confidence"] = "ecard21_only"

    papers_list.append(paper_entry)

# Build ecard-only papers dict
ecard_only_papers = {}
for pe in papers_list:
    if pe["match_confidence"] == "ecard21_only":
        ecard_only_papers[pe["paper_name_ko"]] = {"prices": pe["ecard21"]["prices"]}

# Unmatched bizhows
biz_magnet = []
for item in biz_data:
    if item["category"] == "\uc790\uc11d\uc804\ub2e8\uc9c0":
        biz_magnet.append({
            "product": item["product"],
            "paper": item["paper"],
            "size": item.get("size", ""),
            "qty": item["qty"],
            "price": parse_biz_price(item["price"]),
            "note": "\uc790\uc11d\uc804\ub2e8\uc9c0\ub294 \uc77c\ubc18 \uc804\ub2e8\uc9c0\uc640 \ub2e4\ub978 \uc81c\ud488\uad70",
        })

reference = {
    "description": "2\uc0ac \uc804\ub2e8\uc9c0 \uc6a9\uc9c0 \ub9e4\uce6d \ucc38\uc870 \ub370\uc774\ud130. \uba85\ud568\ucc9c\uad6d(ecard21) vs \ube44\uc988\ud558\uc6b0\uc2a4(bizhows) \uc804\ub2e8\uc9c0 \uac00\uaca9 \ube44\uad50.",
    "updated_at": "2026-03-23",
    "crawled_dates": {"ecard21": "2026-03-23", "bizhows": "2026-03-23"},
    "price_basis": {
        "ecard21": "4000\ub9e4, \uc591\uba74 \uce7c\ub77c, A4 \uae30\uc900",
        "bizhows_\ub300\ub7c9\ud2b9\uac00": "4000\ub9e4, A4 \uae30\uc900 (\ub2e8\uba74/\uc591\uba74 \ubcc4\ub3c4 \uc81c\ud488)",
        "bizhows_\uc18c\ub7c9\uace0\uae09": "50\ub9e4, A4 \uae30\uc900 (\ub2e8\uba74/\uc591\uba74 \ubcc4\ub3c4 \uc81c\ud488)",
        "bizhows_\ud2b9\uc218\uc804\ub2e8\uc9c0": "2000\ub9e4, A4 \uae30\uc900 (\uc591\uba74)",
    },
    "product_types": {
        "\uc77c\ubc18_\ud569\ud310_\uc804\ub2e8\uc9c0": {
            "description": "\uba85\ud568\ucc9c\uad6d \uc77c\ubc18 \ud569\ud310 \uc804\ub2e8\uc9c0 (\uc591\uba74 \uce7c\ub77c, A4, 4000\ub9e4)",
            "ecard21_product": "\uc77c\ubc18 \ud569\ud310 \uc804\ub2e8\uc9c0",
            "bizhows_product": "\ub300\ub7c9 \ud2b9\uac00 \uc804\ub2e8\uc9c0 (\uc591\uba74)",
            "match_type": "product_match",
            "note": "\ube44\uc988\ud558\uc6b0\uc2a4 \ub300\ub7c9\ud2b9\uac00 = \ud569\ud310\uc778\uc1c4 \uc81c\ud488",
        },
        "\uace0\uae09_\ub3c5\ud310_\uc804\ub2e8\uc9c0": {
            "description": "\uba85\ud568\ucc9c\uad6d \uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0 (\uc591\uba74 \uce7c\ub77c, A4, 4000\ub9e4)",
            "ecard21_product": "\uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0",
            "bizhows_product": None,
            "match_type": "ecard21_only",
            "note": "\ube44\uc988\ud558\uc6b0\uc2a4\uc5d0 \ub3c5\ud310 \uc804\ub2e8\uc9c0 \uce74\ud14c\uace0\ub9ac \uc5c6\uc74c",
        },
        "\ubb38\uc5b4\ubc1c_\uc804\ub2e8\uc9c0": {
            "description": "\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0 (\uc591\uba74 \uce7c\ub77c, A4)",
            "ecard21_product": "\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0",
            "bizhows_product": "\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0",
            "match_type": "product_match_qty_diff",
            "qty_diff": "ecard21: 4000\ub9e4 / bizhows: 2000\ub9e4",
            "note": "\uc218\ub7c9\uc774 \ub2e4\ub974\ubbc0\ub85c \uc9c1\uc811 \uac00\uaca9 \ube44\uad50 \uc2dc \uc8fc\uc758 \ud544\uc694",
        },
        "\ubb38\uace0\ub9ac_\uc804\ub2e8\uc9c0": {
            "description": "\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0 (\uc591\uba74 \uce7c\ub77c, A4)",
            "ecard21_product": "\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0",
            "bizhows_product": "\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0",
            "match_type": "product_match_qty_diff",
            "qty_diff": "ecard21: 4000\ub9e4 / bizhows: 2000\ub9e4",
            "note": "\uc218\ub7c9\uc774 \ub2e4\ub974\ubbc0\ub85c \uc9c1\uc811 \uac00\uaca9 \ube44\uad50 \uc2dc \uc8fc\uc758 \ud544\uc694",
        },
        "\ud504\ub9ac\ubbf8\uc5c4_\uc5d0\ucf54_\uc804\ub2e8\uc9c0": {
            "description": "\uba85\ud568\ucc9c\uad6d \ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0",
            "ecard21_product": "\ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0",
            "bizhows_product": None,
            "match_type": "ecard21_only",
        },
        "\ud14c\uc774\ube14_\uc138\ud305\uc9c0": {
            "description": "\uba85\ud568\ucc9c\uad6d \ud14c\uc774\ube14 \uc138\ud305\uc9c0",
            "ecard21_product": "\ud14c\uc774\ube14 \uc138\ud305\uc9c0",
            "bizhows_product": None,
            "match_type": "ecard21_only",
        },
    },
    "papers": papers_list,
    "unmatched": {
        "bizhows_\uc790\uc11d\uc804\ub2e8\uc9c0": biz_magnet,
        "ecard21_only_papers": ecard_only_papers,
    },
    "comparison_summary": {
        "\ub9e4\uce6d\ub41c_\uc6a9\uc9c0_\uc218": matched_count,
        "\uc804\uccb4_\uc6a9\uc9c0_\uc218_ecard21": len(sorted_papers),
        "\uba85\ud568\ucc9c\uad6d_\uace0\uc720": [
            "\uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0 (21\uc885 \uc6a9\uc9c0)",
            "\ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0 (1\uc885)",
            "\ud14c\uc774\ube14 \uc138\ud305\uc9c0 (1\uc885)",
        ],
        "\ube44\uc988\ud558\uc6b0\uc2a4_\uace0\uc720": [
            "\uc18c\ub7c9 \uace0\uae09 \uc804\ub2e8\uc9c0 (50\ub9e4, \uc544\ud2b8\uc9c0/\uc2a4\ub178\uc6b0\uc9c0/\ubaa8\uc870\uc9c0 150g)",
            "\uc885\uc774\uc790\uc11d \uc804\ub2e8\uc9c0 (\ub2e8\uba74/\uc591\uba74, \uac01 8\uc885)",
            "\ud1b5\uc790\uc11d \uc804\ub2e8\uc9c0 (1\uc885)",
        ],
    },
}

with open("C:/Workspace/my-project/analysis/price-monitor/config/flyer_reference.json", "w", encoding="utf-8") as f:
    json.dump(reference, f, ensure_ascii=False, indent=2)
print("flyer_reference.json saved")

# --- Build Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "\uc804\ub2e8\uc9c0 \uac00\uaca9 \ube44\uad50"

header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font = Font(name="Arial", size=10)
blue_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
orange_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
light_blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
light_orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = [
    "\uc6a9\uc9c0",
    "\uc778\uc1c4\ub3c4\uc218",
    "\uba85\ud568\ucc9c\uad6d \uc81c\ud488",
    "\uba85\ud568\ucc9c\uad6d \uc218\ub7c9",
    "\uba85\ud568\ucc9c\uad6d \uac00\uaca9",
    "\ube44\uc988\ud558\uc6b0\uc2a4 \uc6a9\uc9c0",
    "\ube44\uc988\ud558\uc6b0\uc2a4 \uc81c\ud488",
    "\ube44\uc988\ud558\uc6b0\uc2a4 \uc218\ub7c9",
    "\ube44\uc988\ud558\uc6b0\uc2a4 \uac00\uaca9",
    "\ube44\uace0",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = blue_fill if col <= 5 else orange_fill


def find_biz_match(biz_items, paper_type, gram, side, product_keyword):
    for bi in biz_items:
        if (
            bi["paper_type"] == paper_type
            and bi["gram"] == gram
            and bi["side"] == side
            and product_keyword in bi["product"]
        ):
            return bi
    return None


rows = []

# 1. 일반 합판 전단지 (양면) vs 대량특가 (양면)
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\uc77c\ubc18 \ud569\ud310 \uc804\ub2e8\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\uc77c\ubc18 \ud569\ud310 \uc804\ub2e8\uc9c0"]
    biz_match = find_biz_match(biz_a4_items, ptype, gram, "\uc591\uba74", "\ub300\ub7c9")
    note = ""
    if biz_match:
        if biz_match["qty"] != item["quantity"]:
            note = f"\uc218\ub7c9 \ub2e4\ub984 (\uba85\ud568\ucc9c\uad6d {item['quantity']}\ub9e4, \ube44\uc988\ud558\uc6b0\uc2a4 {biz_match['qty']}\ub9e4)"
    else:
        note = "\ube44\uc988\ud558\uc6b0\uc2a4 \ub9e4\uce6d \uc5c6\uc74c"
    rows.append({
        "paper": ecard_paper_display(ptype, gram),
        "side": "\uc591\uba74",
        "ecard_product": "\uc77c\ubc18 \ud569\ud310 \uc804\ub2e8\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": biz_match["paper"] if biz_match else "",
        "biz_product": biz_match["product"] if biz_match else "",
        "biz_qty": biz_match["qty"] if biz_match else "",
        "biz_price": biz_match["price"] if biz_match else "",
        "note": note,
    })

# 2. 대량특가 단면 (bizhows only)
for bi in biz_a4_items:
    if "\ub300\ub7c9" in bi["category"] and bi["side"] == "\ub2e8\uba74":
        rows.append({
            "paper": ecard_paper_display(bi["paper_type"], bi["gram"]),
            "side": "\ub2e8\uba74",
            "ecard_product": "",
            "ecard_qty": "",
            "ecard_price": "",
            "biz_paper": bi["paper"],
            "biz_product": bi["product"],
            "biz_qty": bi["qty"],
            "biz_price": bi["price"],
            "note": "\uba85\ud568\ucc9c\uad6d \uc591\uba74\ub9cc \uc218\uc9d1, \ub2e8\uba74 \ube44\uad50 \ubd88\uac00",
        })

# 3. 고급 독판 전단지 (ecard21 only)
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0"]
    rows.append({
        "paper": ecard_paper_display(ptype, gram),
        "side": "\uc591\uba74",
        "ecard_product": "\uace0\uae09 \ub3c5\ud310 \uc804\ub2e8\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": "",
        "biz_product": "",
        "biz_qty": "",
        "biz_price": "",
        "note": "\ube44\uc988\ud558\uc6b0\uc2a4 \ub3c5\ud310 \uc5c6\uc74c",
    })

# 4. 문어발 전단지
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0"]
    biz_match = find_biz_match(biz_a4_items, ptype, gram, "\uc591\uba74", "\ubb38\uc5b4\ubc1c")
    note = ""
    if biz_match:
        if biz_match["qty"] != item["quantity"]:
            note = f"\uc218\ub7c9 \ub2e4\ub984 (\uba85\ud568\ucc9c\uad6d {item['quantity']}\ub9e4, \ube44\uc988\ud558\uc6b0\uc2a4 {biz_match['qty']}\ub9e4)"
    else:
        note = "\ube44\uc988\ud558\uc6b0\uc2a4 \ub9e4\uce6d \uc5c6\uc74c"
    rows.append({
        "paper": ecard_paper_display(ptype, gram),
        "side": "\uc591\uba74",
        "ecard_product": "\ubb38\uc5b4\ubc1c \uc804\ub2e8\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": biz_match["paper"] if biz_match else "",
        "biz_product": biz_match["product"] if biz_match else "",
        "biz_qty": biz_match["qty"] if biz_match else "",
        "biz_price": biz_match["price"] if biz_match else "",
        "note": note,
    })

# 5. 문고리 전단지
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0"]
    biz_match = find_biz_match(biz_a4_items, ptype, gram, "\uc591\uba74", "\ubb38\uace0\ub9ac")
    note = ""
    if biz_match:
        if biz_match["qty"] != item["quantity"]:
            note = f"\uc218\ub7c9 \ub2e4\ub984 (\uba85\ud568\ucc9c\uad6d {item['quantity']}\ub9e4, \ube44\uc988\ud558\uc6b0\uc2a4 {biz_match['qty']}\ub9e4)"
    else:
        note = "\ube44\uc988\ud558\uc6b0\uc2a4 \ub9e4\uce6d \uc5c6\uc74c"
    rows.append({
        "paper": ecard_paper_display(ptype, gram),
        "side": "\uc591\uba74",
        "ecard_product": "\ubb38\uace0\ub9ac \uc804\ub2e8\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": biz_match["paper"] if biz_match else "",
        "biz_product": biz_match["product"] if biz_match else "",
        "biz_qty": biz_match["qty"] if biz_match else "",
        "biz_price": biz_match["price"] if biz_match else "",
        "note": note,
    })

# 6. 프리미엄 에코 전단지
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0"]
    rows.append({
        "paper": item["paper_name"],
        "side": "\uc591\uba74",
        "ecard_product": "\ud504\ub9ac\ubbf8\uc5c4 \uc5d0\ucf54 \uc804\ub2e8\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": "",
        "biz_product": "",
        "biz_qty": "",
        "biz_price": "",
        "note": "\uba85\ud568\ucc9c\uad6d \uace0\uc720 \uc81c\ud488",
    })

# 7. 테이블 세팅지
for ptype, gram in sorted_papers:
    ecard_prods = ecard_by_paper.get((ptype, gram), {})
    if "\ud14c\uc774\ube14 \uc138\ud305\uc9c0" not in ecard_prods:
        continue
    item = ecard_prods["\ud14c\uc774\ube14 \uc138\ud305\uc9c0"]
    rows.append({
        "paper": item["paper_name"],
        "side": "\uc591\uba74",
        "ecard_product": "\ud14c\uc774\ube14 \uc138\ud305\uc9c0",
        "ecard_qty": item["quantity"],
        "ecard_price": item["price"],
        "biz_paper": "",
        "biz_product": "",
        "biz_qty": "",
        "biz_price": "",
        "note": "\uba85\ud568\ucc9c\uad6d \uace0\uc720 \uc81c\ud488",
    })

# 8. 소량고급 (bizhows only)
for bi in biz_a4_items:
    if "\uc18c\ub7c9" not in bi["category"]:
        continue
    rows.append({
        "paper": ecard_paper_display(bi["paper_type"], bi["gram"]),
        "side": bi["side"],
        "ecard_product": "",
        "ecard_qty": "",
        "ecard_price": "",
        "biz_paper": bi["paper"],
        "biz_product": bi["product"],
        "biz_qty": bi["qty"],
        "biz_price": bi["price"],
        "note": "\ube44\uc988\ud558\uc6b0\uc2a4 \uc18c\ub7c9 \uace0\uc720 (50\ub9e4)",
    })

# Write rows
for r_idx, row in enumerate(rows, 2):
    ws.cell(row=r_idx, column=1, value=row["paper"])
    ws.cell(row=r_idx, column=2, value=row["side"])
    ws.cell(row=r_idx, column=3, value=row["ecard_product"])
    ws.cell(row=r_idx, column=4, value=row["ecard_qty"] if row["ecard_qty"] else "")
    ws.cell(row=r_idx, column=5, value=row["ecard_price"] if row["ecard_price"] else "")
    ws.cell(row=r_idx, column=6, value=row["biz_paper"])
    ws.cell(row=r_idx, column=7, value=row["biz_product"])
    ws.cell(row=r_idx, column=8, value=row["biz_qty"] if row["biz_qty"] else "")
    ws.cell(row=r_idx, column=9, value=row["biz_price"] if row["biz_price"] else "")
    ws.cell(row=r_idx, column=10, value=row["note"])

    is_odd = r_idx % 2 == 0
    for col in range(1, 11):
        cell = ws.cell(row=r_idx, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if col <= 5:
            cell.fill = light_blue_fill if is_odd else white_fill
        else:
            cell.fill = light_orange_fill if is_odd else white_fill
        if col in (4, 5, 8, 9):
            cell.alignment = Alignment(horizontal="right")
            if col in (5, 9) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
        elif col == 2:
            cell.alignment = Alignment(horizontal="center")

col_widths = [22, 8, 18, 14, 14, 30, 26, 14, 14, 44]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A2"

wb.save("C:/Workspace/my-project/analysis/price-monitor/output/flyer_reference_comparison.xlsx")
print(f"Excel saved with {len(rows)} rows")
print(f"Matched papers: {matched_count} / {len(sorted_papers)}")

print("\n--- Bizhows A4 items parsed ---")
for bi in biz_a4_items:
    print(f"  [{bi['category']}] {bi['product']} | {bi['paper']} -> {bi['paper_type']} {bi['gram']}g | {bi['side']} | {bi['qty']}\ub9e4 | {bi['price']}\uc6d0")
