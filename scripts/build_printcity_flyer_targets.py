"""프린트시티 합판전단 엑셀 → config/targets/flyer.yaml printcity 섹션 자동 생성.

입력:
  - data/printcity/flyer2.xlsx (2 paper × 6 size × 2 color × 0.5~100 연)

수집 정책 (2026-05-21~ 결정):
  - 사이즈별 자연 매수 단위(연 정수배)를 표준으로 잡는다.
    A4 만 2개 매수(0.5연 = 2000매, 1연 = 4000매), 나머지는 단일.
  - 1연 = 500매 전지 × 전지당 페이지수.
      A4=8 / A3=4 / A2=2 / B5=16 / B4=8 / B3=4
  - 가격 0 (미판매) / None (카탈로그 없음) 행은 skip.

엑셀 구조:
  R1: header (title, code, quantity, sales, value, calcValue)
  section: MAT:... → SIZ:... → COL:...  data: (quantity=연수, sales=1, value=공급가)
  title 열은 cp949→utf8 mojibake — 라벨은 code-table 로 매핑.
"""
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data/printcity/flyer2.xlsx"
DST = ROOT / "config/targets/flyer.yaml"

# 사이즈별 표준 매수 (사이트 정책 무관 — 비교 기준).
# schemas/flyer.yaml `qty_targets_by_size` 와 동기화 유지.
QTY_TARGETS_BY_SIZE_MAE = {
    "A4": [2000, 4000],   # 0.5연, 1연
    "A3": [2000],         # 1연
    "A2": [4000],         # 4연
    "B5": [8000],         # 1연
    "B4": [4000],         # 1연
    "B3": [8000],         # 4연
}

# 1연 매수 = 500매 전지 × 전지당 페이지수.
SIZE_PAGES_PER_SHEET = {"A4": 8, "A3": 4, "A2": 2, "B5": 16, "B4": 8, "B3": 4}

PAPER_CODE_TO_NAME = {
    "MAT:ART-90": "아트-90g",
    "MAT:MOJ-80": "모조-80g",
}

SIZE_CODE_TO_INFO = {
    "SIZ:LEJ-A4": ("A4", "A4 국8절(210x297)"),
    "SIZ:LEJ-A3": ("A3", "A3 국4절(297x420)"),
    "SIZ:LEJ-A2": ("A2", "A2 국2절(420x597)"),
    "SIZ:LEJ-B5": ("B5", "B5 16절(182x257)"),
    "SIZ:LEJ-B4": ("B4", "B4 8절(257x367)"),
    "SIZ:LEJ-B3": ("B3", "B3 4절(367x517)"),
}

COLOR_CODE_TO_NAME = {
    "COL:40": "단면4도",
    "COL:44": "양면8도",
}


def mae_to_reams(size: str, mae: int) -> float:
    return mae / (500 * SIZE_PAGES_PER_SHEET[size])


def parse_xlsx(path: Path) -> dict:
    """엑셀 → (paper_code, size_code, color_code, reams) → price."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    prices = {}
    paper_code = size_code = color_code = None
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 2).value
        qty = ws.cell(r, 3).value
        val = ws.cell(r, 5).value
        if isinstance(code, str) and code.startswith("MAT:"):
            paper_code = code; continue
        if isinstance(code, str) and code.startswith("SIZ:"):
            size_code = code; continue
        if isinstance(code, str) and code.startswith("COL:"):
            color_code = code; continue
        if not isinstance(qty, (int, float)):
            continue
        if paper_code and size_code and color_code:
            prices[(paper_code, size_code, color_code, float(qty))] = val
    return prices


def build_items(prices: dict) -> list:
    # size → size_code 역색인
    size_to_code = {info[0]: code for code, info in SIZE_CODE_TO_INFO.items()}

    items = []
    skipped = []
    for size, targets_mae in QTY_TARGETS_BY_SIZE_MAE.items():
        size_code = size_to_code[size]
        _, size_label = SIZE_CODE_TO_INFO[size_code]
        for paper_code, paper_name in PAPER_CODE_TO_NAME.items():
            for color_code, color_name in COLOR_CODE_TO_NAME.items():
                for target_mae in targets_mae:
                    reams = mae_to_reams(size, target_mae)
                    price = prices.get((paper_code, size_code, color_code, reams))
                    if not price:
                        skipped.append((paper_name, size, color_name, target_mae,
                                        "미판매(0)" if price == 0 else "카탈로그 없음"))
                        continue
                    items.append({
                        "product": "합판전단",
                        "paper": paper_name,
                        "paper_code": paper_code,
                        "color_mode": color_name,
                        "color_code": color_code,
                        "size": size,
                        "size_code": size_code,
                        "size_label": size_label,
                        "qty": int(target_mae),
                        "qty_yeon": reams,
                        "price": int(price),
                    })
    return items, skipped


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    if not SRC.exists():
        print(f"❌ {SRC} 없음")
        sys.exit(1)

    prices = parse_xlsx(SRC)
    items, skipped = build_items(prices)

    # 기존 yaml 의 다른 사이트 섹션은 보존.
    data = {}
    if DST.exists():
        data = yaml.safe_load(DST.read_text(encoding="utf-8")) or {}

    pc_section = {
        "_description": "정적 엑셀 source — data/printcity/flyer2.xlsx (사이즈별 연 단위 가격표)",
        "sources": ["data/printcity/flyer2.xlsx"],
        "filters_applied": {
            "qty_targets_by_size": QTY_TARGETS_BY_SIZE_MAE,
            "paper": list(PAPER_CODE_TO_NAME.values()),
            "color": list(COLOR_CODE_TO_NAME.values()),
        },
        "price_vat_included": False,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "items": items,
    }
    # printcity 키를 맨 앞으로
    ordered = {"printcity": pc_section}
    for k, v in data.items():
        if k != "printcity":
            ordered[k] = v

    DST.write_text(
        yaml.safe_dump(ordered, allow_unicode=True, sort_keys=False,
                       default_flow_style=False, width=140),
        encoding="utf-8",
    )

    from collections import Counter
    by_size = Counter(it["size"] for it in items)
    by_paper = Counter(it["paper"] for it in items)
    print(f"✅ {DST}: printcity 섹션 갱신")
    print(f"  총 {len(items)} items")
    print(f"  size 분포: {dict(by_size)}")
    print(f"  paper 분포: {dict(by_paper)}")
    if skipped:
        print(f"  skip {len(skipped)} 건:")
        for p, sz, c, q, why in skipped:
            print(f"    - {p} × {sz} × {c} × {q}매 — {why}")


if __name__ == "__main__":
    main()
