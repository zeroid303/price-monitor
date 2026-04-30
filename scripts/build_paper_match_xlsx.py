"""raw 카드 명함 (offset/digital) 의 paper_name × site × coating → canonical 매칭 후
 xlsx 로 저장. user 검토용.

매칭 기준:
- 경쟁사 가격 모니터링 피드백.xlsx 의 사이트별 alias 우선 매칭
- 평량 ±20g 룰 (위반 시 다른 canonical 분리 — 메모 표시)
- 영어 prefix → 한글 (Extra → 엑스트라)
- coating: 정규화 (비코팅/무광코팅/유광코팅/벨벳코팅/홀로그램코팅)
- 플라스틱 (μ paper) / 단종 paper 는 매칭 제외 (메모)
- offset/digital 분리 출력
- 단일 사이트 등장 OK
- 헷갈림 → 메모 컬럼에 사소한 것도 적기
"""
import json
import os
import re
import sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FEEDBACK = "C:/Users/Admin/Downloads/경쟁사 가격 모니터링 피드백.xlsx"
OUT_PATH = os.path.expanduser("~/Downloads/명함_용지_매칭결과_2026-04-30.xlsx")

# 단종 paper (피드백 시트에 명시) — schema 매칭 제외
DISCONTINUED = {
    "린넨펄", "엔틱골드펄", "포레스트그린", "Extra 엠보 270g",
    "엔틱골드펄 240g", "포레스트그린 216g",
}

# 영어 prefix 한글 변환
PREFIX_MAP = [
    ("Extra ", "엑스트라 "),
    ("extra ", "엑스트라 "),
]


def normalize_canonical_name(name: str) -> str:
    s = name
    for k, v in PREFIX_MAP:
        if s.startswith(k):
            s = v + s[len(k):]
    return s.strip()


# coating 정규화 (raw → canonical)
COATING_MAP = {
    None: "비코팅",
    "": "비코팅",
    "코팅없음": "비코팅",
    "무코팅": "비코팅",
    "비코팅": "비코팅",
    "(무코팅)": "비코팅",
    "(비코팅)": "비코팅",
    "무광코팅": "무광코팅",
    "양면무광코팅": "무광코팅",
    "단면(무광) 코팅": "무광코팅",
    "양면(무광) 코팅": "무광코팅",
    "(무광코팅)": "무광코팅",
    "유광코팅": "유광코팅",
    "양면유광코팅": "유광코팅",
    "단면(유광) 코팅": "유광코팅",
    "양면(유광) 코팅": "유광코팅",
    "(유광코팅)": "유광코팅",
    "벨벳코팅": "벨벳코팅",
    "양면벨벳코팅": "벨벳코팅",
    "(벨벳코팅)": "벨벳코팅",
    "홀로그램코팅 도트-양면": "홀로그램코팅(도트)",
    "홀로그램코팅 심플-양면": "홀로그램코팅(심플)",
}

# 사이트별 raw paper_name 안에 박힌 coating 패턴
COATING_INLINE_RE = re.compile(r"\((무광코팅|유광코팅|벨벳코팅|무코팅|비코팅|UV유광코팅|UV코팅)\)")
# printcity dash 패턴 ("스노우화이트 216g- 양면무광코팅")
COATING_DASH_RE = re.compile(r"-\s*(양면|단면)?(무광|유광|벨벳)코팅")


def detect_coating(paper_name: str, coating_field: str | None) -> tuple[str, str]:
    """(coating_canonical, source) — paper_name 에 박혀있으면 우선, 없으면 coating field."""
    # 1) paper_name 괄호
    m = COATING_INLINE_RE.search(paper_name)
    if m:
        tok = m.group(1)
        if "무코팅" in tok or "비코팅" in tok:
            return ("비코팅", "paper_name")
        if "무광" in tok:
            return ("무광코팅", "paper_name")
        if "유광" in tok and "UV" in tok:
            return ("유광코팅", "paper_name")  # UV유광 → 유광 통합
        if "유광" in tok:
            return ("유광코팅", "paper_name")
        if "벨벳" in tok:
            return ("벨벳코팅", "paper_name")
    # 2) printcity dash
    m = COATING_DASH_RE.search(paper_name)
    if m:
        kind = m.group(2)
        return ({"무광": "무광코팅", "유광": "유광코팅", "벨벳": "벨벳코팅"}.get(kind, "비코팅"),
                "paper_name")
    # 3) coating field
    if coating_field is not None:
        normalized = COATING_MAP.get(coating_field, None)
        if normalized:
            return (normalized, "coating_field")
        # printcity 양면무광/유광 등
        if "무광" in coating_field:
            return ("무광코팅", "coating_field")
        if "유광" in coating_field:
            return ("유광코팅", "coating_field")
        if "벨벳" in coating_field:
            return ("벨벳코팅", "coating_field")
        if "홀로그램" in coating_field:
            return ("홀로그램코팅", "coating_field")
        if "코팅없음" in coating_field or coating_field.strip() in ("", "None"):
            return ("비코팅", "coating_field")
    # 4) fallback
    return ("비코팅(default)", "default")


# 평량 추출
WEIGHT_RE = re.compile(r"(\d{2,4})\s*g(?!/)")  # 250g (g/㎡ 의 g 는 별도 처리)
WEIGHT_GMS_RE = re.compile(r"(\d{2,4})\s*g\s*/\s*㎡")  # adsland digital "300 g/㎡"
WEIGHT_MICRON_RE = re.compile(r"(\d{2,4})\s*[μu](?!nique)")  # 250μ, 250u


def extract_weight(paper_name: str) -> tuple[int | str | None, bool]:
    """(weight, is_plastic). is_plastic 은 μ 단위면 True."""
    if not paper_name:
        return (None, False)
    m = WEIGHT_GMS_RE.search(paper_name)
    if m:
        return (int(m.group(1)), False)
    m = WEIGHT_RE.search(paper_name)
    if m:
        return (int(m.group(1)), False)
    m = WEIGHT_MICRON_RE.search(paper_name)
    if m:
        return (int(m.group(1)), True)
    return (None, False)


# ── 피드백 xlsx 파싱 ──

def load_feedback() -> list[dict]:
    """피드백 xlsx 의 행을 list[{canonical, weight_range, sites, memo}] 로 반환."""
    wb = load_workbook(FEEDBACK, data_only=True)
    ws = wb["명함 용지"]
    rows = []
    for r in range(2, ws.max_row + 1):
        canon = ws.cell(r, 1).value
        gram = ws.cell(r, 2).value
        if canon is None and gram is None:
            continue
        canon = (canon or "").strip()
        # 사이트별 alias (C=printcity, D=swadpia, E=wowpress, F=adsland, G=dtpia)
        sites = {
            "printcity": _split_aliases(ws.cell(r, 3).value),
            "swadpia":   _split_aliases(ws.cell(r, 4).value),
            "wowpress":  _split_aliases(ws.cell(r, 5).value),
            "adsland":   _split_aliases(ws.cell(r, 6).value),
            "dtpia":     _split_aliases(ws.cell(r, 7).value),
        }
        memo = " | ".join(filter(None, [_s(ws.cell(r, 8).value), _s(ws.cell(r, 9).value)]))
        rows.append({
            "row": r,
            "canonical": canon,
            "weight_range": _s(gram),
            "sites": sites,
            "memo": memo,
        })
    return rows


def _s(v) -> str:
    if v is None: return ""
    return str(v).strip()


def _split_aliases(v) -> list[str]:
    if v is None: return []
    s = str(v)
    parts = [p.strip() for p in s.split("/")]
    return [p for p in parts if p]


# ── 매칭 ──

def build_reverse_map(feedback: list[dict]) -> dict:
    """site → alias_lower → row.

    alias 와 alias 의 (평량 strip 변형) 둘 다 등록.
    """
    rev = defaultdict(dict)
    for row in feedback:
        for site, aliases in row["sites"].items():
            for a in aliases:
                rev[site][a.strip()] = row
                # 평량 suffix 제거 변형도 등록 (예: "스노우지 250g" → "스노우지")
                strip = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", a).strip()
                strip = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", strip).strip()
                if strip and strip != a:
                    rev[site].setdefault(strip, row)
    return rev


def _normalize_paper_for_match(site: str, paper_name: str) -> list[str]:
    """매칭 시도용 변형 list 생성 (긴 → 짧은)."""
    cands = [paper_name]
    pn = paper_name

    # adsland digital: "고급지 반누보 227 g/㎡" → "반누보 227g" 또는 "반누보화이트 227g"
    if site == "adsland":
        # "PaperSort X Y g/㎡" 패턴
        m = re.match(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+(.+?)\s+(\d{2,4})\s*g\s*/\s*㎡\s*$", pn)
        if m:
            paper_part, weight = m.group(2), m.group(3)
            cands.append(f"{paper_part} {weight}g")
            cands.append(paper_part)

    # printcity dash 패턴: "스노우화이트 250g-양면무광,비코팅" → "스노우화이트 250g"
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?[, ].*$", "", pn).strip()
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?\s*$", "", pn2).strip()
    if pn2 != pn:
        cands.append(pn2)

    # printcity dash: "스노우화이트 250g" → "스노우화이트-250g" 형식 또는 역
    pn3 = re.sub(r"-(\d{2,4}g)", r" \1", pn)
    if pn3 != pn:
        cands.append(pn3)

    # 코팅 괄호 제거: "스노우지(무광코팅) 250g" → "스노우지 250g"
    pn4 = COATING_INLINE_RE.sub("", pn).replace("  ", " ").strip()
    if pn4 != pn:
        cands.append(pn4)

    # 평량 suffix 제거 (마지막 fallback)
    base = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", pn).strip()
    base = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", base).strip()
    if base and base != pn:
        cands.append(base)

    # adsland 의 paperSort 만 제거한 base
    if site == "adsland":
        m = re.match(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+(.+?)\s+(\d{2,4})\s*g\s*/\s*㎡\s*$", pn)
        if m:
            cands.append(m.group(2))

    # 중복 제거
    seen = set(); out = []
    for c in cands:
        if c not in seen:
            seen.add(c); out.append(c)
    return out


def match_canonical(site: str, paper_name: str, weight: int | None,
                    rev: dict, feedback: list[dict]) -> tuple[str, str, list[str]]:
    """(canonical, match_method, memos)."""
    memos = []
    site_map = rev.get(site, {})

    candidates = _normalize_paper_for_match(site, paper_name)

    # 1. 같은 사이트 alias 매칭 (변형들 시도)
    for cand in candidates:
        if cand in site_map:
            row = site_map[cand]
            canonical = normalize_canonical_name(row["canonical"])
            wmemos = check_weight_in_range(weight, row["weight_range"], canonical)
            memos.extend(wmemos)
            if cand != paper_name:
                memos.append(f"변형 매칭: {cand!r}")
            if row["memo"]: memos.append(f"피드백비고: {row['memo'][:80]}")
            return (canonical, "exact" if cand == paper_name else "variant", memos)

    # 2. cross-site (다른 사이트의 alias) 매칭 — 변형 모두 시도
    for cand in candidates:
        for fb_row in feedback:
            for s, aliases in fb_row["sites"].items():
                if cand in aliases:
                    canonical = normalize_canonical_name(fb_row["canonical"])
                    wmemos = check_weight_in_range(weight, fb_row["weight_range"], canonical)
                    memos.extend(wmemos)
                    memos.append(f"cross-site({s} alias: {cand!r})")
                    return (canonical, "cross", memos)

    # 3. canonical 명 substring 매칭 (예: "반누보화이트" → "반누보 ...")
    for cand in candidates:
        for fb_row in feedback:
            cn = fb_row["canonical"].strip()
            if not cn or len(cn) < 3: continue
            # canonical 명이 cand 안에 들어가면 잠재 매칭
            base_cand = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", cand).strip()
            if cn in base_cand or base_cand in cn:
                canonical = normalize_canonical_name(cn)
                wmemos = check_weight_in_range(weight, fb_row["weight_range"], canonical)
                memos.extend(wmemos)
                memos.append(f"canonical-substring({cn!r} ↔ {cand!r}) — 추정 (검토 필요)")
                return (canonical, "substring", memos)

    # 4. 미매칭
    memos.append("⚠ 피드백 xlsx 에 없음 — canonical 신규 등록 필요")
    return ("(미매칭)", "miss", memos)


def check_weight_in_range(weight: int | None, range_str: str, canonical: str) -> list[str]:
    """평량이 피드백 시트의 범위 안에 있는지. 위반 시 메모."""
    memos = []
    if weight is None or not range_str:
        return memos
    # range_str 예: "250", "200~219", "186~250", "(평량없음)"
    rs = range_str.replace("(평량없음)", "").strip()
    if not rs: return memos
    if "~" in rs:
        try:
            lo, hi = rs.split("~", 1)
            lo, hi = int(float(lo.strip())), int(float(hi.strip()))
            # ±20g 룰: range 외라도 hi+20 이내까지는 OK (대표값 가까움)
            if not (lo - 20 <= weight <= hi + 20):
                memos.append(f"⚠ 평량 {weight}g 가 {canonical}({rs}g) 범위 ±20g 벗어남 — 다른 canonical 분리 필요")
        except ValueError:
            pass
    else:
        try:
            target = int(float(rs))
            if abs(weight - target) > 20:
                memos.append(f"⚠ 평량 {weight}g 가 {canonical}({target}g) ±20g 벗어남 — 다른 canonical 분리 필요")
        except ValueError:
            pass
    return memos


# ── raw record 집계 + 매칭 ──

def aggregate_raw_per_category(category: str) -> list[dict]:
    """(site, paper_name) unique 그룹. coating list 는 셀에 합쳐서 표시."""
    groups = defaultdict(lambda: {"count": 0, "products": set(),
                                   "coatings": set()})
    for fn in sorted(os.listdir(os.path.join(ROOT, "output"))):
        if not fn.endswith("_raw_now.json"): continue
        if f"_{category}_" not in fn: continue
        site = fn.split("_")[0]
        d = json.load(open(os.path.join(ROOT, "output", fn), encoding="utf-8"))
        for i in d.get("items", []):
            pn = i.get("paper_name") or ""
            coating = i.get("coating") or "(none)"
            key = (site, pn)
            groups[key]["count"] += 1
            groups[key]["coatings"].add(coating)
            if i.get("product"):
                groups[key]["products"].add(i["product"])
    return [
        {"site": k[0], "paper_name": k[1],
         "coatings": sorted(v["coatings"]),
         "count": v["count"], "products": sorted(v["products"])}
        for k, v in groups.items()
    ]


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    feedback = load_feedback()
    rev = build_reverse_map(feedback)
    print(f"피드백 행: {len(feedback)}")

    wb = Workbook()
    for cat, sheet_name in (("card_offset", "card_offset"), ("card_digital", "card_digital")):
        records = aggregate_raw_per_category(cat)
        records.sort(key=lambda r: (r["site"], r["paper_name"]))
        if cat == "card_offset":
            ws = wb.active; ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append([
            "사이트", "paper_name(raw)", "coating(raw, distinct)", "레코드 수", "제품(들)",
            "평량g(추출)", "재질(플라스틱)", "코팅(정규화-주)",
            "대표용지(canonical)", "canonical_key", "매칭 방법",
            "메모(헷갈림 / 결정 사항)",
        ])
        miss = 0
        viol = 0
        for r in records:
            site = r["site"]; pn = r["paper_name"]
            coatings = r["coatings"]
            coat_str = " / ".join(coatings)
            # 첫 coating 으로 정규화 시도 (대표값)
            primary_coat = coatings[0] if coatings else None
            if primary_coat == "(none)":
                primary_coat = None
            weight, is_plastic = extract_weight(pn)
            coat_norm, coat_src = detect_coating(pn, primary_coat)
            memos = []
            if is_plastic:
                memos.append(f"플라스틱 재질({weight}μ) — schema 매칭 제외 권장")
                canonical = "(플라스틱-제외)"
                method = "skip-plastic"
            else:
                canonical, method, m_memos = match_canonical(site, pn, weight, rev, feedback)
                memos.extend(m_memos)
                if method == "miss":
                    miss += 1
                if any("벗어남" in m for m in memos):
                    viol += 1
            if canonical and canonical not in ("(미매칭)", "(플라스틱-제외)") and weight is not None:
                key = f"{canonical}_{weight}g_{coat_norm}"
            else:
                key = "(미정)"
            # coating 다양 표시
            if len(coatings) > 1:
                memos.append(f"coating raw 다양({len(coatings)}종) — 정규화 후 별 record 로 분리됨")
            elif coat_src == "default":
                memos.append("coating 출처: default(비코팅 추정)")
            ws.append([
                site, pn, coat_str, r["count"], " / ".join(r["products"]),
                weight if weight is not None else "",
                "Y" if is_plastic else "",
                coat_norm, canonical, key, method,
                " | ".join(memos),
            ])
        # 스타일
        hf = Font(bold=True, color="FFFFFF"); hp = PatternFill("solid", fgColor="4472C4")
        ha = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = hf; cell.fill = hp; cell.alignment = ha
        widths = {1: 10, 2: 30, 3: 18, 4: 8, 5: 25, 6: 10, 7: 10,
                  8: 14, 9: 25, 10: 30, 11: 10, 12: 60}
        for col_idx, w in widths.items():
            ws.column_dimensions[chr(64 + col_idx)].width = w
        ws.freeze_panes = "A2"
        # 미매칭 행 빨간색
        red = PatternFill("solid", fgColor="FFCCCC")
        yellow = PatternFill("solid", fgColor="FFF2CC")
        for r in range(2, ws.max_row + 1):
            method = ws.cell(r, 11).value
            memo = ws.cell(r, 12).value or ""
            if method == "miss":
                for c in range(1, 13):
                    ws.cell(r, c).fill = red
            elif "벗어남" in memo or method == "skip-plastic":
                for c in range(1, 13):
                    ws.cell(r, c).fill = yellow
        print(f"  {sheet_name}: rows={len(records)} miss={miss} viol={viol}")

    wb.save(OUT_PATH)
    print(f"\n✅ saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
