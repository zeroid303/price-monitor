import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "봉투 가격 비교"

# Styles
font_default = Font(name="Arial", size=10)
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
fill_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
fill_row_alt1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
fill_row_alt2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# Headers
headers = [
    "용지", "봉투 종류", "명함천국 수량", "명함천국 가격",
    "비즈하우스 용지", "비즈하우스 수량", "비즈하우스 가격",
    "가격차(비즈-명함)", "비고"
]
header_fills = [fill_blue, fill_blue, fill_blue, fill_blue,
                fill_orange, fill_orange, fill_orange,
                fill_gray, fill_gray]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border
    cell.fill = header_fills[col - 1]

# Data: (용지, 봉투종류, 명함천국수량, 명함천국가격, 비즈하우스용지, 비즈하우스수량, 비즈하우스가격, 비고)
data = [
    # === 매칭 가능한 용지 (컬러 500매) ===
    ("모조지 100g", "소봉투 (일반)", "500매(컬러)", 39000, "모조지,백색,100g", "500매", 37600, ""),
    ("모조지 100g", "소봉투 (자켓)", "500매(컬러)", 39000, "모조지,백색,100g", "500매", 37600, ""),
    ("모조지 120g", "대봉투 (A4)", "500매(컬러)", 88000, "모조지,백색,120g", "500매", 104800, ""),
    ("모조지 120g", "소봉투 (일반)", "500매(컬러)", 43500, "모조지,백색,120g", "500매", 43200, ""),
    ("모조지 120g", "소봉투 (자켓)", "500매(컬러)", 43500, "모조지,백색,120g", "500매", 43200, ""),
    ("레자크체크백색 110g", "대봉투 (A4)", "500매(컬러)", 93000, "체크무늬 레자크지,백색,110g", "500매", 113500, "비즈하우스 품절"),
    ("레자크체크백색 110g", "소봉투 (일반)", "500매(컬러)", 50000, "체크무늬 레자크지,백색,110g", "500매", 47700, "비즈하우스 품절"),
    ("레자크체크백색 110g", "소봉투 (자켓)", "500매(컬러)", 50000, "체크무늬 레자크지,백색,110g", "500매", 47700, "비즈하우스 품절"),
    ("레자크줄무늬 110g", "대봉투 (A4)", "500매(컬러)", 93000, "줄무늬 레자크지,백색,110g", "500매", 113500, "비즈하우스 품절"),
    ("레자크줄무늬 110g", "소봉투 (일반)", "500매(컬러)", 50000, "줄무늬 레자크지,백색,110g", "500매", 47700, "비즈하우스 품절"),
    ("레자크줄무늬 110g", "소봉투 (자켓)", "500매(컬러)", 50000, "줄무늬 레자크지,백색,110g", "500매", 47700, "비즈하우스 품절"),
    # === 비즈하우스에만 있는 용지 ===
    ("모조지 150g", "대봉투 (A4)", "-", None, "모조지,백색,150g", "500매", 280000, "명함천국에 해당 용지 없음"),
    # === 명함천국만 있는 용지 (주요 컬러) ===
    ("모조지 120g", "중봉투 (6절)", "500매(컬러)", 148700, "-", "-", None, "비즈하우스에 중봉투 없음"),
    ("모조지 120g", "중봉투 (9절)", "500매(컬러)", 148700, "-", "-", None, "비즈하우스에 중봉투 없음"),
    ("모조지 180g", "대봉투 (A4)", "500매(컬러)", 201100, "-", "-", None, "비즈하우스에 180g 없음"),
    ("모조지 180g", "소봉투 (일반)", "500매(컬러)", 185600, "-", "-", None, "비즈하우스에 180g 없음"),
    ("레자크 120g", "대봉투 (A4)", "500매(컬러)", 184300, "-", "-", None, "비즈하우스에 레자크120g 없음"),
    ("레자크 120g", "소봉투 (일반)", "500매(컬러)", 161300, "-", "-", None, "비즈하우스에 레자크120g 없음"),
    ("크라프트 98g", "대봉투 (A4)", "500매(컬러)", 149500, "-", "-", None, "비즈하우스에 크라프트봉투 없음"),
    ("크라프트 98g", "중봉투 (6절)", "500매(컬러)", 149500, "-", "-", None, "비즈하우스에 크라프트봉투 없음"),
    ("화일지 120g", "대봉투 (A4)", "500매(컬러)", 163900, "-", "-", None, "명함천국 고유"),
    ("화일지 120g", "소봉투 (일반)", "500매(컬러)", 148100, "-", "-", None, "명함천국 고유"),
    ("스타펄 120g", "대봉투 (A4)", "500매(컬러)", 414700, "-", "-", None, "명함천국 고유 (프리미엄)"),
    ("스타펄 120g", "소봉투 (일반)", "500매(컬러)", 334100, "-", "-", None, "명함천국 고유 (프리미엄)"),
    ("탄트 120g", "대봉투 (A4)", "500매(컬러)", 209500, "-", "-", None, "명함천국 고유"),
    ("탄트 120g", "소봉투 (일반)", "500매(컬러)", 178100, "-", "-", None, "명함천국 고유"),
    ("밍크 120g", "대봉투 (A4)", "500매(컬러)", 181900, "-", "-", None, "명함천국 고유"),
    ("레이드 120g", "대봉투 (A4)", "500매(컬러)", 191500, "-", "-", None, "명함천국 고유"),
    ("랑데뷰 105g", "대봉투 (A4)", "500매(컬러)", 239300, "-", "-", None, "명함천국 고유"),
    ("랑데뷰 130g", "대봉투 (A4)", "500매(컬러)", 270300, "-", "-", None, "명함천국 고유"),
    ("랑데뷰 160g", "대봉투 (A4)", "500매(컬러)", 288700, "-", "-", None, "명함천국 고유"),
    # === 흑백 봉투 (비즈하우스에 흑백 없음) ===
    ("모조지 120g", "대봉투 (A4)", "1000매(흑백)", 89000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크줄무늬 120g", "대봉투 (A4)", "1000매(흑백)", 110000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크체크백색 120g", "대봉투 (A4)", "1000매(흑백)", 110000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("화일지 120g", "대봉투 (A4)", "1000매(흑백)", 109000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("크라프트 98g", "대봉투 (A4)", "1000매(흑백)", 78000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("모조지 100g", "중봉투 (6절)", "1000매(흑백)", 83000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    ("모조지 100g", "중봉투 (9절)", "1000매(흑백)", 74000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    ("모조지 100g", "소봉투 (일반)", "1000매(흑백)", 43500, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("모조지 100g", "소봉투 (자켓)", "1000매(흑백)", 48500, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("크라프트 98g", "중봉투 (6절)", "1000매(흑백)", 69000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    ("크라프트 98g", "중봉투 (9절)", "1000매(흑백)", 62000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    ("레자크줄무늬 120g", "소봉투 (일반)", "1000매(흑백)", 58000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크체크백색 120g", "소봉투 (일반)", "1000매(흑백)", 58000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크줄무늬 120g", "소봉투 (자켓)", "1000매(흑백)", 61000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크체크백색 120g", "소봉투 (자켓)", "1000매(흑백)", 61000, "-", "-", None, "비즈하우스에 흑백봉투 없음"),
    ("레자크줄무늬 120g", "중봉투 (6절)", "1000매(흑백)", 100000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    ("레자크줄무늬 120g", "중봉투 (9절)", "1000매(흑백)", 87000, "-", "-", None, "비즈하우스에 중봉투/흑백 없음"),
    # === 비즈하우스 소량 전용 (500매 없음) ===
    ("-", "대봉투 (A4)", "-", None, "반투명 A4 대봉투 (180g)", "10매만", None, "비즈하우스 소량10매 전용 (38,680원/10매). 500매 없음"),
    ("-", "대봉투 (A4)", "-", None, "트레싱지 대봉투", "10매만", None, "비즈하우스 소량10매 전용 (25,440원/10매). 500매 없음"),
]

for i, row_data in enumerate(data):
    row = i + 2
    paper, size, ec_qty, ec_price, bh_paper, bh_qty, bh_price, note = row_data

    # Calculate price diff
    diff = None
    if ec_price is not None and bh_price is not None:
        diff = bh_price - ec_price

    ws.cell(row=row, column=1, value=paper)
    ws.cell(row=row, column=2, value=size)
    ws.cell(row=row, column=3, value=ec_qty)
    ws.cell(row=row, column=4, value=ec_price if ec_price is not None else "-")
    ws.cell(row=row, column=5, value=bh_paper)
    ws.cell(row=row, column=6, value=bh_qty)
    ws.cell(row=row, column=7, value=bh_price if bh_price is not None else "-")
    ws.cell(row=row, column=8, value=diff if diff is not None else "-")
    ws.cell(row=row, column=9, value=note)

    # Apply styles
    alt_fill = fill_row_alt1 if i % 2 == 0 else fill_row_alt2
    for col in range(1, 10):
        cell = ws.cell(row=row, column=col)
        cell.font = font_default
        cell.border = thin_border
        cell.fill = alt_fill
        if col in (4, 7, 8):
            cell.alignment = align_right
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
        elif col in (1, 5, 9):
            cell.alignment = align_left
        else:
            cell.alignment = align_center

# Column widths
col_widths = [22, 16, 14, 14, 32, 12, 14, 18, 42]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze pane
ws.freeze_panes = "A2"

output_path = "C:/Workspace/my-project/analysis/price-monitor/output/envelope_reference_comparison.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
