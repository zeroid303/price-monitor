# -*- coding: utf-8 -*-
"""전단지 2사 가격 비교 매칭 레퍼런스 생성 스크립트"""
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent

# ── Load data ──
with open(BASE / "output/ecard21_flyer_prices.json", "r", encoding="utf-8") as f:
    ecard = json.load(f)
with open(BASE / "output/bizhows_flyer_prices.json", "r", encoding="utf-8") as f:
    bizhows = json.load(f)


def parse_biz_price(price_str):
    """할인가 있으면 할인가, 없으면 원가 반환"""
    if not price_str:
        return None
    m = re.findall(r"([\d,]+)원", price_str)
    if len(m) >= 2:
        return int(m[-1].replace(",", ""))
    elif len(m) == 1:
        return int(m[0].replace(",", ""))
    return None


# ── Build bizhows lookup ──
biz_muneo = {}
biz_mungori = {}
biz_dae_danmyeon = []
biz_dae_yangmyeon = []
biz_soryang = []
biz_jasoek = []

for item in bizhows:
    cat = item["category"]
    prod = item["product"]
    paper = item["paper"]
    qty = item["qty"]
    price_raw = item["price"]
    price = parse_biz_price(price_raw)
    side = item.get("side", "")

    if prod == "문어발 전단지":
        biz_muneo[paper] = {"price": price, "price_raw": price_raw, "qty": qty, "side": side, "url": item.get("url", "")}
    elif prod == "문고리 전단지":
        biz_mungori[paper] = {"price": price, "price_raw": price_raw, "qty": qty, "side": side, "url": item.get("url", "")}
    elif "대량" in cat and "단면" in prod and str(qty) == "4000":
        biz_dae_danmyeon.append({"paper": paper, "price": price, "price_raw": price_raw, "qty": qty, "url": item.get("url", "")})
    elif "대량" in cat and "양면" in prod and str(qty) == "4000":
        biz_dae_yangmyeon.append({"paper": paper, "price": price, "price_raw": price_raw, "qty": qty, "url": item.get("url", "")})
    elif "소량" in cat:
        biz_soryang.append(item)
    elif "자석" in cat:
        biz_jasoek.append(item)

# ── Paper matching (bizhows -> ecard21) ──
paper_match = {
    "아트지 100g": "아트지 100g",
    "아트지 120g": "아트지 120g",
    "스노우지 100g": "스노우화이트 100g",
    "스노우지 120g": "스노우화이트 120g",
    "모조지 100g": "모조지 100g",
    "모조지 120g": "모조지 120g",
}

# ── ecard21 lookups ──
ecard_muneo = {}
for item in ecard["products"].get("문어발 전단지", []):
    if item["size_code"] == "pssj5":
        ecard_muneo[item["paper_name"]] = {"price": item["price"], "paper_code": item["paper_code"]}

ecard_mungori = {}
for item in ecard["products"].get("문고리 전단지", []):
    if item["size_code"] == "pssq4":
        ecard_mungori[item["paper_name"]] = {"price": item["price"], "paper_code": item["paper_code"]}

ecard_hapan = {}
for item in ecard["products"].get("일반 합판 전단지", []):
    ecard_hapan[item["paper_name"]] = {"price": item["price"], "paper_code": item["paper_code"]}

ecard_dokpan = {}
for item in ecard["products"].get("고급 독판 전단지", []):
    ecard_dokpan[item["paper_name"]] = {"price": item["price"], "paper_code": item["paper_code"]}

ecard_eco = {}
for item in ecard["products"].get("프리미엄 에코 전단지", []):
    ecard_eco[item["paper_name"]] = {"price": item["price"], "paper_code": item["paper_code"]}

# ══════════════════════════════════════════════════════════
# ── Build reference JSON ──
# ══════════════════════════════════════════════════════════
reference = {
    "description": "2사 전단지 용지 매칭 참조 데이터. 명함천국(ecard21) vs 비즈하우스(bizhows) 전단지 가격 비교.",
    "updated_at": "2026-03-23",
    "crawled_dates": {"ecard21": "2026-03-23", "bizhows": "2026-03-23"},
    "price_basis": {
        "ecard21": "4000매, 양면 칼라, A4 기준",
        "bizhows_대량특가": "4000매, A4 기준 (단면/양면 별도 제품). 용지명 미확인(seq).",
        "bizhows_특수전단지": "2000매, 양면, A4 기준 (4000매 미지원)",
    },
    "product_types": {
        "일반_합판_전단지": {
            "description": "명함천국 일반 합판 전단지 (양면 칼라, A4, 4000매)",
            "ecard21_product": "일반 합판 전단지",
            "bizhows_product": "대량 특가 전단지 (양면)",
            "match_type": "product_match",
            "note": "비즈하우스 대량특가 = 합판인쇄 제품. 용지명이 seq 형태로 수집되어 정확한 용지 매칭 불가.",
        },
        "고급_독판_전단지": {
            "description": "명함천국 고급 독판 전단지 (양면 칼라, A4, 4000매)",
            "ecard21_product": "고급 독판 전단지",
            "bizhows_product": None,
            "match_type": "ecard21_only",
            "note": "비즈하우스에 독판 전단지 카테고리 없음",
        },
        "문어발_전단지": {
            "description": "문어발 전단지 (양면 칼라, A4)",
            "ecard21_product": "문어발 전단지",
            "bizhows_product": "문어발 전단지",
            "match_type": "product_match_qty_diff",
            "qty_diff": "ecard21: 4000매 / bizhows: 2000매",
            "note": "수량이 다르므로 직접 가격 비교 시 주의 필요",
        },
        "문고리_전단지": {
            "description": "문고리 전단지 (양면 칼라, A4)",
            "ecard21_product": "문고리 전단지",
            "bizhows_product": "문고리 전단지",
            "match_type": "product_match_qty_diff",
            "qty_diff": "ecard21: 4000매 / bizhows: 2000매",
            "note": "수량이 다르므로 직접 가격 비교 시 주의 필요",
        },
        "프리미엄_에코_전단지": {
            "description": "명함천국 프리미엄 에코 전단지",
            "ecard21_product": "프리미엄 에코 전단지",
            "bizhows_product": None,
            "match_type": "ecard21_only",
        },
        "테이블_세팅지": {
            "description": "명함천국 테이블 세팅지",
            "ecard21_product": "테이블 세팅지",
            "bizhows_product": None,
            "match_type": "ecard21_only",
        },
    },
    "papers": [],
    "unmatched": {
        "bizhows_대량특가_양면_4000": [],
        "bizhows_대량특가_단면_4000": [],
        "bizhows_소량고급": [],
        "bizhows_자석전단지": [],
        "ecard21_only_papers": {},
    },
    "comparison_summary": {},
}

# ── Matched papers ──
for biz_paper, ecard_paper in paper_match.items():
    ecard_prices = {}
    biz_prices = {}

    muneo_e = ecard_muneo.get(ecard_paper)
    if muneo_e:
        ecard_prices["문어발_양면_4000"] = muneo_e["price"]

    dokpan_e = ecard_dokpan.get(ecard_paper)
    if dokpan_e:
        ecard_prices["고급독판_양면_4000"] = dokpan_e["price"]

    mungori_e = ecard_mungori.get(ecard_paper)
    if mungori_e:
        ecard_prices["문고리_양면_4000"] = mungori_e["price"]

    muneo_b = biz_muneo.get(biz_paper)
    if muneo_b:
        biz_prices["문어발_양면_2000"] = muneo_b["price"]

    mungori_b = biz_mungori.get(biz_paper)
    if mungori_b:
        biz_prices["문고리_양면_2000"] = mungori_b["price"]

    reference["papers"].append({
        "paper_id": ecard_paper.replace(" ", "_").lower(),
        "paper_name_ko": ecard_paper,
        "match_confidence": "high",
        "ecard21": {
            "paper_name": ecard_paper,
            "paper_code": (muneo_e or dokpan_e or {}).get("paper_code", ""),
            "prices": ecard_prices,
        },
        "bizhows": {
            "paper_name": biz_paper,
            "prices": biz_prices,
            "note": "비즈하우스 문어발/문고리는 2000매 기준 (4000매 미지원)",
        },
    })

# ── Unmatched: ecard21 only papers ──
ecard21_only = {}
for source, items in [
    ("문어발_양면_4000", ecard_muneo),
    ("문고리_양면_4000", ecard_mungori),
    ("고급독판_양면_4000", ecard_dokpan),
    ("일반합판_양면_4000", ecard_hapan),
    ("에코_양면_4000", ecard_eco),
]:
    for pn, info in items.items():
        if pn not in paper_match.values():
            ecard21_only.setdefault(pn, {"prices": {}})
            ecard21_only[pn]["prices"][source] = info["price"]

reference["unmatched"]["ecard21_only_papers"] = ecard21_only

# ── Unmatched: bizhows ──
for item in biz_dae_yangmyeon:
    reference["unmatched"]["bizhows_대량특가_양면_4000"].append({
        "paper_seq": item["paper"],
        "price": item["price"],
        "price_raw": item["price_raw"],
        "qty": item["qty"],
        "note": "용지명 미확인 (seq). 합판인쇄 제품으로 명함천국 일반합판에 대응 추정. 크롤러 재실행 필요.",
        "url": item["url"],
    })

for item in biz_dae_danmyeon:
    reference["unmatched"]["bizhows_대량특가_단면_4000"].append({
        "paper_seq": item["paper"],
        "price": item["price"],
        "price_raw": item["price_raw"],
        "qty": item["qty"],
        "note": "용지명 미확인 (seq). 명함천국은 양면 칼라 기준으로 수집되어 단면 비교 불가.",
        "url": item["url"],
    })

for item in biz_soryang:
    reference["unmatched"]["bizhows_소량고급"].append({
        "product": item["product"],
        "paper": item["paper"],
        "qty": item["qty"],
        "price": parse_biz_price(item["price"]),
        "note": "50매 소량 제품. 4000매 기준 비교 불가.",
    })

for item in biz_jasoek:
    reference["unmatched"]["bizhows_자석전단지"].append({
        "product": item["product"],
        "paper": item["paper"],
        "size": item["size"],
        "qty": item["qty"],
        "price": parse_biz_price(item["price"]),
        "note": "자석전단지는 일반 전단지와 다른 제품군",
    })

reference["comparison_summary"] = {
    "매칭된_용지_수": len(paper_match),
    "매칭_가능_제품": {
        "문어발_전단지": "6종 용지 매칭 (수량 다름: ecard21 4000매 vs bizhows 2000매)",
        "문고리_전단지": "6종 용지 매칭 (수량 다름: ecard21 4000매 vs bizhows 2000매)",
    },
    "비교_제한사항": {
        "수량_불일치": "비즈하우스 문어발/문고리 = 2000매, 명함천국 = 4000매",
        "대량특가_용지_미확인": "비즈하우스 대량특가의 용지명이 seq 형태로 수집됨. 크롤러 재실행 또는 수동 확인 필요.",
        "단면_미수집": "명함천국은 양면 칼라 기준 수집. 비즈하우스 대량특가 단면과 비교 불가.",
    },
    "명함천국_고유": [
        "고급 독판 전단지 (21종 용지)",
        "프리미엄 에코 전단지 (1종)",
        "테이블 세팅지 (1종)",
    ],
    "비즈하우스_고유": [
        "소량 고급 전단지 (50매, 3종)",
        "종이자석 전단지 (단면/양면, 각 8종)",
        "통자석 전단지 (1종)",
    ],
}

with open(BASE / "config/flyer_reference.json", "w", encoding="utf-8") as f:
    json.dump(reference, f, ensure_ascii=False, indent=2)
print("config/flyer_reference.json saved")

# ══════════════════════════════════════════════════════════
# ── Build Excel ──
# ══════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "전단지 가격비교"

hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
hdr_blue = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
hdr_orange = PatternFill(start_color="E07020", end_color="E07020", fill_type="solid")
body_font = Font(name="Arial", size=10)
alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
wht = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
bdr = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

headers = [
    "용지", "인쇄도수", "사이즈",
    "명함천국 제품", "명함천국 가격",
    "비즈하우스 용지", "비즈하우스 가격",
    "비고",
]

for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.border = bdr
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = hdr_blue if ci <= 5 else hdr_orange

rows = []

# ── Section A: 문어발 전단지 (매칭됨, 수량 다름) ──
for biz_p, ecard_p in paper_match.items():
    me = ecard_muneo.get(ecard_p)
    mb = biz_muneo.get(biz_p)
    rows.append([
        ecard_p, "양면", "A4",
        "문어발 전단지 (4000매)", me["price"] if me else None,
        biz_p + " [2000매]", mb["price"] if mb else None,
        "수량 다름 (ecard21 4000매 vs bizhows 2000매)",
    ])

# ── Section B: 문고리 전단지 (매칭됨, 수량 다름) ──
for biz_p, ecard_p in paper_match.items():
    me = ecard_mungori.get(ecard_p)
    mb = biz_mungori.get(biz_p)
    if not me:
        continue
    rows.append([
        ecard_p, "양면", "A4",
        "문고리 전단지 (4000매)", me["price"],
        biz_p + " [2000매]" if mb else "", mb["price"] if mb else None,
        "수량 다름 (ecard21 4000매 vs bizhows 2000매)" if mb else "비즈하우스 해당 용지 없음",
    ])

# ── Section C: 일반 합판 전단지 (매칭 불가) ──
for pn, info in ecard_hapan.items():
    rows.append([
        pn, "양면", "A4",
        "일반 합판 전단지 (4000매)", info["price"],
        "", None,
        "비즈하우스 대량특가 용지명 미확인 (seq). 매칭 불가.",
    ])

# ── Section D: 고급 독판 전단지 (ecard21 only) ──
for pn, info in ecard_dokpan.items():
    if pn in paper_match.values():
        continue
    rows.append([
        pn, "양면", "A4",
        "고급 독판 전단지 (4000매)", info["price"],
        "", None,
        "비즈하우스에 독판 전단지 없음",
    ])

# ── Section E: 프리미엄 에코 (ecard21 only) ──
for pn, info in ecard_eco.items():
    rows.append([
        pn, "양면", "A4",
        "프리미엄 에코 전단지 (4000매)", info["price"],
        "", None,
        "명함천국 고유 제품",
    ])

# ── Section F: 비즈하우스 대량특가 양면 (unmatched) ──
for item in biz_dae_yangmyeon:
    rows.append([
        "", "양면", "A4",
        "", None,
        item["paper"] + " [4000매]", item["price"],
        "대량특가 양면. 용지명 미확인 (seq). 명함천국 일반합판 대응 추정.",
    ])

# ── Section G: 비즈하우스 대량특가 단면 (unmatched) ──
for item in biz_dae_danmyeon:
    rows.append([
        "", "단면", "A4",
        "", None,
        item["paper"] + " [4000매]", item["price"],
        "대량특가 단면. 용지명 미확인 (seq). 명함천국 단면 미수집.",
    ])

# ── Write rows ──
for ri, row in enumerate(rows, 2):
    fill = alt if ri % 2 == 0 else wht
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = body_font
        c.fill = fill
        c.border = bdr
        if ci in (5, 7) and val is not None:
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")
        elif ci in (2, 3):
            c.alignment = Alignment(horizontal="center")

# ── Column widths ──
widths = [22, 10, 8, 26, 14, 24, 14, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

xlsx_path = BASE / "output/flyer_reference_comparison.xlsx"
wb.save(xlsx_path)
print(f"output/flyer_reference_comparison.xlsx saved ({len(rows)} rows)")
