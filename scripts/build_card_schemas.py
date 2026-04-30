"""raw 기반 + 피드백 xlsx 기반 schema 자동 생성.

산출:
  config/schemas/card_offset.yaml
  config/schemas/card_digital.yaml

매칭 키: (대표용지, 평량g, 코팅) — coating 은 별도 axis 로 정규화되므로
schema 의 canonical 단위는 (paper, weight_g) 페어. aliases 는 사이트별 raw paper_name.

룰:
- 평량 ±20g 매칭. 위반 시 동일 paper 의 새 weight_g 행 자동 추가.
- 피드백 xlsx 에 있는 paper 는 alias 그대로 등록.
- 신규 paper (raw 에만 등장) 는 auto-canonical 등록.
- 영어 prefix → 한글 (Extra → 엑스트라).
- 플라스틱 (μ) — 유포지 외 제외.
- 단종 paper (린넨펄/엔틱골드펄/포레스트그린/Extra 엠보 270g) — schema 미포함.
- offset / digital 분리.
"""
import json
import os
import re
import sys
from collections import defaultdict

import yaml
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FEEDBACK = "C:/Users/Admin/Downloads/경쟁사 가격 모니터링 피드백.xlsx"

DISCONTINUED = {"린넨펄", "엔틱골드펄", "포레스트그린", "Extra 엠보 270g", "Extra 엠보"}


def normalize_canonical_name(name: str) -> str:
    s = name.strip()
    s = re.sub(r"^Extra\s+", "엑스트라 ", s)
    s = re.sub(r"^extra\s+", "엑스트라 ", s)
    return s


# 부가 표기 strip — 매칭 시 (FSC)/(D)/(신상품)/당일판가능 무시
AUX_TOKEN_RES = [
    re.compile(r"\s*\(FSC\)\s*"),
    re.compile(r"\s*\(D\)\s*"),
    re.compile(r"\s*\(신상품(-?[가-힣A-Za-z]*)?\)\s*"),
    re.compile(r"\s*-?\s*당일판가능\s*"),
    re.compile(r"\s+"),  # 다중 공백 정리
]


def strip_aux_tokens(s: str) -> str:
    """raw paper_name 의 부가 표기 (FSC)/(D)/-당일판가능 등 제거. 매칭/canonical 정리용."""
    if not s: return s
    out = s
    for pat in AUX_TOKEN_RES[:-1]:
        out = pat.sub(" ", out)
    out = AUX_TOKEN_RES[-1].sub(" ", out).strip()
    return out


# ── 정규식 ──
COATING_INLINE_RE = re.compile(r"\((무광코팅|유광코팅|벨벳코팅|무코팅|비코팅|UV유광코팅|UV코팅)\)")
COATING_DASH_RE = re.compile(r"-\s*(양면|단면)?(무광|유광|벨벳)코팅")
WEIGHT_GMS_RE = re.compile(r"(\d{2,4})\s*g\s*/\s*㎡")
WEIGHT_RE = re.compile(r"(\d{2,4})\s*g(?!/)")
WEIGHT_MICRON_RE = re.compile(r"(\d{2,4})\s*[μuU](?!nique)")


def detect_coating(paper_name, coating_field):
    pn = paper_name or ""
    cf = (coating_field or "").strip() if coating_field else ""
    m = COATING_INLINE_RE.search(pn)
    if m:
        tok = m.group(1)
        if "무코팅" in tok or "비코팅" in tok: return "비코팅"
        if "무광" in tok: return "무광코팅"
        if "유광" in tok: return "유광코팅"
        if "벨벳" in tok: return "벨벳코팅"
    m = COATING_DASH_RE.search(pn)
    if m:
        return {"무광": "무광코팅", "유광": "유광코팅", "벨벳": "벨벳코팅"}.get(m.group(2), "비코팅")
    if cf:
        if "코팅없음" in cf or cf in ("무코팅", "비코팅"): return "비코팅"
        if "무광" in cf: return "무광코팅"
        if "유광" in cf: return "유광코팅"
        if "벨벳" in cf: return "벨벳코팅"
        if "홀로그램" in cf: return "홀로그램코팅"
    return "비코팅"


def extract_weight(paper_name, paper_weight_text=None):
    """(weight_int, is_plastic_micron). paper_name 에 없으면 paper_weight_text fallback."""
    for src in (paper_name, paper_weight_text):
        if not src: continue
        m = WEIGHT_GMS_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_MICRON_RE.search(src)
        if m: return (int(m.group(1)), True)
    return (None, False)


# ── 피드백 xlsx 파싱 ──

def load_feedback():
    wb = load_workbook(FEEDBACK, data_only=True)
    ws = wb["명함 용지"]
    rows = []
    for r in range(2, ws.max_row + 1):
        canon = ws.cell(r, 1).value
        gram = ws.cell(r, 2).value
        if canon is None and gram is None: continue
        canon = (canon or "").strip()
        if not canon or canon in DISCONTINUED: continue
        sites = {
            "printcity": _split_aliases(ws.cell(r, 3).value),
            "swadpia":   _split_aliases(ws.cell(r, 4).value),
            "wowpress":  _split_aliases(ws.cell(r, 5).value),
            "adsland":   _split_aliases(ws.cell(r, 6).value),
            "dtpia":     _split_aliases(ws.cell(r, 7).value),
        }
        memo = " | ".join(filter(None, [_s(ws.cell(r, 8).value), _s(ws.cell(r, 9).value)]))
        rows.append({
            "canonical": canon,
            "weight_range": _s(gram),
            "sites": sites,
            "memo": memo,
        })
    return rows


def _s(v): return "" if v is None else str(v).strip()


def _split_aliases(v):
    if v is None: return []
    return [p.strip() for p in str(v).split("/") if p.strip()]


def feedback_weight_int(weight_range: str):
    """피드백 시트의 평량 표기 → 대표 평량 int. 범위면 평균/대표 hi 사용."""
    rs = weight_range.replace("(평량없음)", "").strip()
    if not rs: return None
    try:
        if "~" in rs:
            lo, hi = [int(float(x.strip())) for x in rs.split("~", 1)]
            return (lo + hi) // 2
        return int(float(rs))
    except ValueError:
        return None


def feedback_weight_range(weight_range: str):
    """피드백 시트의 평량 표기 → (lo, hi) tuple. 범위 없으면 (w, w)."""
    rs = weight_range.replace("(평량없음)", "").strip()
    if not rs: return None
    try:
        if "~" in rs:
            lo, hi = [int(float(x.strip())) for x in rs.split("~", 1)]
            return (lo, hi)
        w = int(float(rs))
        return (w, w)
    except ValueError:
        return None


# ── 매칭 ──

def build_reverse_map(feedback):
    rev = defaultdict(dict)  # site → alias → fb_row. 부가표기 strip 변형도 등록.
    for row in feedback:
        for site, aliases in row["sites"].items():
            for a in aliases:
                rev[site][a] = row
                # 부가 표기 strip 변형
                stripped = strip_aux_tokens(a)
                if stripped != a:
                    rev[site].setdefault(stripped, row)
                # 평량 suffix 제거
                base = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", a).strip()
                base = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", base).strip()
                if base and base != a:
                    rev[site].setdefault(base, row)
                # base + aux strip 변형
                base_stripped = strip_aux_tokens(base) if base else ""
                if base_stripped and base_stripped != base:
                    rev[site].setdefault(base_stripped, row)
    return rev


def normalize_for_match(site, paper_name):
    cands = [paper_name]
    pn = paper_name
    # 부가 표기 strip
    pn_stripped = strip_aux_tokens(pn)
    if pn_stripped != pn: cands.append(pn_stripped)
    # adsland paperSort
    for src in (pn, pn_stripped):
        if site == "adsland":
            m = re.match(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+(.+?)\s+(\d{2,4})\s*g\s*/\s*㎡\s*$", src)
            if m:
                cands.append(f"{m.group(2)} {m.group(3)}g")
                cands.append(m.group(2))
    # printcity dash 코팅 strip
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?[, ].*$", "", pn).strip()
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?\s*$", "", pn2).strip()
    if pn2 != pn: cands.append(pn2)
    pn3 = re.sub(r"-(\d{2,4}g)", r" \1", pn)
    if pn3 != pn: cands.append(pn3)
    pn4 = COATING_INLINE_RE.sub("", pn).replace("  ", " ").strip()
    if pn4 != pn: cands.append(pn4)
    # 평량 suffix strip
    for src in (pn, pn_stripped):
        base = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", src).strip()
        base = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", base).strip()
        if base and base != src: cands.append(base)
    seen = set(); out = []
    for c in cands:
        if c not in seen:
            seen.add(c); out.append(c)
    return out


def match_canonical(site, paper_name, weight, rev, feedback):
    """(canonical_name, fb_row | None, viol_bool)."""
    cands = normalize_for_match(site, paper_name)
    site_map = rev.get(site, {})
    for cand in cands:
        if cand in site_map:
            row = site_map[cand]
            wr = feedback_weight_range(row["weight_range"])
            viol = False
            if wr and weight is not None:
                lo, hi = wr
                viol = not (lo - 20 <= weight <= hi + 20)
            return (normalize_canonical_name(row["canonical"]), row, viol)
    for cand in cands:
        for fb_row in feedback:
            for s, aliases in fb_row["sites"].items():
                if cand in aliases:
                    wr = feedback_weight_range(fb_row["weight_range"])
                    viol = False
                    if wr and weight is not None:
                        lo, hi = wr
                        viol = not (lo - 20 <= weight <= hi + 20)
                    return (normalize_canonical_name(fb_row["canonical"]), fb_row, viol)
    return (None, None, False)


def auto_canonical_name(paper_name: str) -> str:
    """미매칭 raw paper_name → auto canonical 이름.

    예: "고급지 띤또레또 순백색 250 g/㎡" → "띤또레또 순백색"
        "매쉬멜로우 209g (FSC)" → "매쉬멜로우"
        "스타드림 골드 240g (FSC)" → "스타드림 골드"
    """
    s = paper_name
    # adsland paperSort 제거
    s = re.sub(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+", "", s)
    # 평량 제거
    s = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", s).strip()
    s = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", s).strip()
    # 코팅 괄호 제거
    s = COATING_INLINE_RE.sub("", s).strip()
    # 부가 표기 strip
    s = strip_aux_tokens(s)
    # 트리밍 + trailing dash/underscore 제거
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip("-_ ").strip()
    return normalize_canonical_name(s)


# ── 집계 ──

SITES = ["printcity", "swadpia", "wowpress", "adsland", "dtpia"]


def aggregate_canonical(category, feedback, rev):
    """canonical → weight → site → set(aliases). + memo set.

    raw → site/weight 매칭 후, 평량 클러스터링:
      - 같은 사이트가 한 canonical 에서 여러 weight 가지면 → 분리 유지
      - 그 외 ±20g 안 평량들은 한 weight 행으로 통합 (대표 평량 = 최빈/중앙값)
    """
    raw_records = []  # (canonical, weight, site, paper_name, viol, memo)

    for fn in sorted(os.listdir(os.path.join(ROOT, "output"))):
        if not fn.endswith("_raw_now.json"): continue
        if f"_{category}_" not in fn: continue
        site = fn.split("_")[0]
        d = json.load(open(os.path.join(ROOT, "output", fn), encoding="utf-8"))
        for i in d.get("items", []):
            pn = i.get("paper_name") or ""
            pwt = i.get("paper_weight_text") or ""
            weight, is_plastic = extract_weight(pn, pwt)
            if is_plastic and "유포" not in pn:
                continue
            w = weight if weight is not None else 0
            canonical, fb_row, viol = match_canonical(site, pn, weight, rev, feedback)
            if canonical is None:
                canonical = auto_canonical_name(pn)
                if not canonical:
                    continue
                memo = "auto-canonical (피드백 X)"
            else:
                memo = ""
                if viol:
                    memo = f"평량 ±20g 위반 (피드백 범위: {fb_row['weight_range']}) — 분리됨"
                if fb_row and fb_row["memo"]:
                    memo = (memo + " | " + fb_row["memo"][:80]) if memo else fb_row["memo"][:80]
            # canonical 표준화 — known alias map 적용 (raw 단계에서 적용해야 cluster 정확)
            canonical = apply_canonical_alias(canonical)
            # Extra/엑스트라 prefix 가 raw paper_name 에 있으면 canonical 명도 분기
            if ("Extra" in pn or "엑스트라" in pn) and not canonical.startswith("엑스트라 "):
                canonical = f"엑스트라 {canonical}"
            raw_records.append((canonical, w, site, pn, viol, memo))

    # 평량 클러스터링 (per canonical)
    table = defaultdict(lambda: defaultdict(lambda: {
        "sites": defaultdict(set),
        "memos": set(),
    }))
    by_canon = defaultdict(list)
    for r in raw_records:
        by_canon[r[0]].append(r)

    for canonical, recs in by_canon.items():
        # 1. 모든 평량 cluster 화 (±20g)
        weights = sorted({r[1] for r in recs if r[1] > 0})
        clusters = cluster_weights(weights, tolerance=20)

        # 2. 각 cluster 별로 "같은 사이트가 cluster 안 여러 weight 가지는지" 체크
        for cluster in clusters:
            site_w_in_cluster = defaultdict(set)
            for canonical_, w, site, pn, viol, memo in recs:
                if w in cluster:
                    site_w_in_cluster[site].add(w)
            same_site_multi_in_cluster = any(len(ws) > 1 for ws in site_w_in_cluster.values())

            if same_site_multi_in_cluster:
                # 분리 유지 (같은 사이트 같은 제품에서 여러 평량 = 다른 paper)
                for canonical_, w, site, pn, viol, memo in recs:
                    if w in cluster:
                        table[canonical][w]["sites"][site].add(display_alias(pn))
                        if memo: table[canonical][w]["memos"].add(memo)
            else:
                # 통합 (다른 사이트끼리 평량 차이 = 같은 paper)
                rep_w = int(round(sum(cluster) / len(cluster)))
                for canonical_, w, site, pn, viol, memo in recs:
                    if w in cluster:
                        table[canonical][rep_w]["sites"][site].add(display_alias(pn))
                        if memo: table[canonical][rep_w]["memos"].add(memo)
                        if w != rep_w:
                            table[canonical][rep_w]["memos"].add(
                                f"{site} 실제 평량 {w}g (대표 {rep_w}g 로 통합 — ±20g 안)"
                            )
        # 3. weight 0 (평량 추출 실패) 케이스 별도
        for canonical_, w, site, pn, viol, memo in recs:
            if w == 0:
                table[canonical][0]["sites"][site].add(display_alias(pn))
                if memo: table[canonical][0]["memos"].add(memo)

    # 4. post-process: Extra/엑스트라 → 별 canonical 분리
    table = post_process_extra(table)
    # 5. post-process: canonical 이름 통합 (alias map + 공백 차이)
    table = post_process_merge_aliases(table)
    return table


def display_alias(alias: str) -> str:
    """Schema 표시용 alias 정규화 — adsland paperSort prefix 제거 + 'g/㎡' → 'g'."""
    s = re.sub(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+", "", alias)
    s = re.sub(r"(\d+)\s*g\s*/\s*㎡", r"\1g", s)
    return s.strip()


# 알려진 canonical 통합 매핑 — auto_canonical 결과를 표준 canonical 로 정규화
CANONICAL_ALIAS_MAP = {
    "반누보": "반누보화이트",
    "썬샤인": "썬샤인화이트",
    "션샤인": "썬샤인화이트",
    "시리오 진주": "시리오 스노우",
    "시리오펄(진주)": "시리오펄 진주",
    "스타드림골드": "스타드림 골드",
    "스타드림실버": "스타드림 실버",
    "스타드림쿼츠": "스타드림 수정",  # 쿼츠 = 수정 (피드백 R47)
    "스타드림 쿼츠": "스타드림 수정",
    "스타드림오팔": "스타드림 오팔",
    "랑데뷰 내츄럴": "랑데뷰 네츄럴",
    "랑데뷰네츄럴": "랑데뷰 네츄럴",
    "마쉬멜로우": "머쉬멜로우",
    "매쉬멜로우": "머쉬멜로우",
    "머시멜로우": "머쉬멜로우",
    "마쉬멜로우 미색": "머쉬멜로우",
    "마쉬멜로우 화이트": "머쉬멜로우",
    "스코틀랜드": "스코트랜드",
    "스코트랜드": "스코트랜드",  # 통일
    "샤이니": "사파이어펄",
    "샤이니 골드펄": "사파이어펄",
    "메탈릭 실버": "스타드림 실버",
    "메탈릭 화이트": "기본 화이트카드",
    "스타드림 화이트": "기본 화이트카드",
    "메탈아이스": "기본 화이트카드",
    "마제스틱": "키칼라골드",
    "마제스틱(마블화이트)": "키칼라골드",
}


def apply_canonical_alias(name: str) -> str:
    return CANONICAL_ALIAS_MAP.get(name, name)


def post_process_merge_aliases(table):
    """canonical 이름 통합:
       1) CANONICAL_ALIAS_MAP 의 매핑 적용
       2) 공백 차이만 다른 canonical 통합 ('스타드림골드' → '스타드림 골드' 형태로 합칠 때 더 긴 것 우선)
    """
    # 1) alias map 적용
    new_table = defaultdict(lambda: defaultdict(lambda: {
        "sites": defaultdict(set), "memos": set(),
    }))
    for canonical, weights in table.items():
        target = apply_canonical_alias(canonical)
        for w, info in weights.items():
            for site, aliases in info["sites"].items():
                new_table[target][w]["sites"][site].update(aliases)
            new_table[target][w]["memos"].update(info["memos"])

    # 2) 공백 정규화 통합
    canonicals = list(new_table.keys())
    norm_groups = defaultdict(list)
    for c in canonicals:
        norm_groups[c.replace(" ", "")].append(c)

    for nname, members in norm_groups.items():
        if len(members) < 2: continue
        std = max(members, key=lambda x: x.count(" "))
        for m in members:
            if m == std: continue
            for w, info in new_table[m].items():
                for site, aliases in info["sites"].items():
                    new_table[std][w]["sites"][site].update(aliases)
                new_table[std][w]["memos"].update(info["memos"])
            del new_table[m]

    return new_table


def post_process_extra(table):
    """모든 alias 가 Extra/엑스트라 prefix 가지면 별도 canonical '엑스트라 X' 로 분리."""
    new_table = defaultdict(lambda: defaultdict(lambda: {
        "sites": defaultdict(set), "memos": set(),
    }))
    for canonical, weights in table.items():
        for w, info in weights.items():
            all_aliases = []
            for site, aliases in info["sites"].items():
                all_aliases.extend(aliases)
            if not all_aliases:
                new_table[canonical][w]["sites"] = info["sites"]
                new_table[canonical][w]["memos"] = info["memos"]
                continue
            is_extra = lambda a: ("Extra" in a or "엑스트라" in a)
            if all(is_extra(a) for a in all_aliases):
                new_canon = canonical if canonical.startswith("엑스트라 ") else f"엑스트라 {canonical}"
                for site, aliases in info["sites"].items():
                    new_table[new_canon][w]["sites"][site].update(aliases)
                new_table[new_canon][w]["memos"].update(info["memos"])
            else:
                new_table[canonical][w]["sites"] = info["sites"]
                new_table[canonical][w]["memos"] = info["memos"]
    return new_table


def cluster_weights(weights: list[int], tolerance: int = 20) -> list[list[int]]:
    """정렬된 평량 list 를 ±tolerance 안 그룹으로 클러스터링.

    어떤 두 평량이라도 cluster 안에서 max-min ≤ tolerance.
    """
    if not weights: return []
    weights = sorted(weights)
    clusters = []
    cur = [weights[0]]
    for w in weights[1:]:
        if w - cur[0] <= tolerance:
            cur.append(w)
        else:
            clusters.append(cur)
            cur = [w]
    clusters.append(cur)
    return clusters


# ── schema yaml 생성 ──

def build_schema_yaml(table, category_label, weight_tolerance=20):
    paper_canonical = {}
    sorted_canonicals = sorted(table.keys())
    for canonical in sorted_canonicals:
        if canonical in DISCONTINUED: continue
        weight_dict = table[canonical]
        weight_keys = sorted(weight_dict.keys())
        # weight 별 entry
        weights_info = {}
        all_aliases = defaultdict(list)
        for w in weight_keys:
            sites_aliases = weight_dict[w]["sites"]
            site_dict = {}
            for site in SITES:
                aliases = sorted(sites_aliases.get(site, set()))
                if aliases:
                    site_dict[site] = aliases
                    for a in aliases:
                        all_aliases[site].append(a)
            entry = {"aliases": site_dict}
            memos = sorted(m for m in weight_dict[w]["memos"] if m)
            if memos:
                entry["memo"] = " | ".join(memos)
            weights_info[w if w > 0 else "(평량없음)"] = entry

        paper_canonical[canonical] = {
            "weights": weights_info,
            "all_weights_g": [w for w in weight_keys if w > 0],
        }

    schema = {
        "_description": (
            f"카드 명함 ({category_label}) 정규화 룰. "
            "canonical key = (paper, weight_g, coating). "
            "raw paper_name → 사이트별 alias 매칭으로 canonical 결정. "
            "평량 ±20g 위반 시 동일 paper 안에서 별 weight 행으로 분리."
        ),
        "_match_axes": {
            "axes": ["paper_name", "paper_weight_g", "coating", "print_mode", "size", "qty"],
            "weight_tolerance_g": weight_tolerance,
        },
        "_normalization": {
            "paper_name": {
                "_description": (
                    "용지명 정규화. canonical paper-level + weight-level 분기. "
                    "각 weight entry 의 aliases 가 사이트별 raw paper_name."
                ),
                "canonical": paper_canonical,
                "noise_suffix_regex": r"(\d+)\s*g.*$",
            },
            "coating": {
                "_description": "coating 정규화. raw → 비코팅/무광코팅/유광코팅/벨벳코팅/홀로그램코팅",
                "default": "비코팅",
                "allowed": ["비코팅", "무광코팅", "유광코팅", "벨벳코팅", "홀로그램코팅"],
                "aliases": {
                    "비코팅": ["코팅없음", "무코팅", "(무코팅)", "(비코팅)", None],
                    "무광코팅": ["양면무광코팅", "단면(무광) 코팅", "양면(무광) 코팅", "(무광코팅)"],
                    "유광코팅": ["양면유광코팅", "단면(유광) 코팅", "양면(유광) 코팅", "(유광코팅)", "UV유광코팅", "UV코팅"],
                    "벨벳코팅": ["양면벨벳코팅", "(벨벳코팅)"],
                    "홀로그램코팅": ["홀로그램코팅 도트-양면", "홀로그램코팅 심플-양면"],
                },
            },
            "print_mode": {
                "_description": "도수/색도 정규화. 기본 양면칼라.",
                "default": "양면칼라",
                "allowed": ["단면칼라", "양면칼라", "단면별색", "양면별색", "백색", "별색"],
                "aliases": {
                    "단면칼라": ["단면4도", "단면 4도 (컬러)", "단면 칼라4도", "4/0", "단면4도 (인쇄)"],
                    "양면칼라": ["양면8도", "양면 칼라8도", "4/4", "앞면 4도 (컬러) / 뒷면 4도 (컬러)"],
                    "단면별색": ["단면 별색"],
                    "양면별색": ["양면 별색"],
                    "백색": ["앞면 1도 (백색) / 뒷면 1도 (백색)"],
                },
            },
            "size": {
                "_description": "사이즈 정규화 — '90x50' 표준",
                "default": "90x50",
                "format": "WxH",
                "regex": r"(\d+)\s*[x×*X]\s*(\d+)",
            },
            "qty": {
                "_description": "수량 정규화 — int 변환",
                "default": 200,
            },
        },
    }
    return schema


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    feedback = load_feedback()
    rev = build_reverse_map(feedback)
    print(f"피드백 행: {len(feedback)}")

    for cat, label in (("card_offset", "오프셋"), ("card_digital", "디지털")):
        table = aggregate_canonical(cat, feedback, rev)
        schema = build_schema_yaml(table, label)
        out_path = os.path.join(ROOT, f"config/schemas/{cat}.yaml")
        with open(out_path, "w", encoding="utf-8") as f:
            yaml.safe_dump(schema, f, allow_unicode=True, sort_keys=False,
                           default_flow_style=False, width=200)
        n_canonical = len(table)
        n_weights = sum(len(v) for v in table.values())
        n_aliases = sum(
            len(s) for v in table.values() for w in v.values() for s in w["sites"].values()
        )
        print(f"  {cat}: canonical={n_canonical} weight_entries={n_weights} aliases={n_aliases}")
        print(f"    → {out_path}")


if __name__ == "__main__":
    main()
