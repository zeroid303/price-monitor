"""raw 의 (canonical, 평량g, 코팅) 매칭 결과를 피드백 xlsx 와 같은 표 형식으로 출력.

각 행 = (대표용지, 평량g, 코팅) canonical key.
컬럼 = 프린트시티 / 성원 / 와우 / 애즈 / 디티 (사이트 paper_name 들).
시트 분리: card_offset / card_digital / 미매칭 (신규 canonical 후보).

빈 셀은 그 사이트 미수집 또는 미매칭. 평량 ±20g 위반 행은 노란색.
"""
import json
import os
import re
import sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FEEDBACK = "C:/Users/Admin/Downloads/경쟁사 가격 모니터링 피드백.xlsx"
OUT_PATH = os.path.expanduser("~/Downloads/명함_용지_매칭표_2026-04-30.xlsx")

# 영어 prefix → 한글
PREFIX_MAP = [("Extra ", "엑스트라 "), ("extra ", "엑스트라 ")]


def normalize_canonical_name(name: str) -> str:
    s = name
    for k, v in PREFIX_MAP:
        if s.startswith(k):
            s = v + s[len(k):]
    return s.strip()


# coating 정규화
COATING_INLINE_RE = re.compile(r"\((무광코팅|유광코팅|벨벳코팅|무코팅|비코팅|UV유광코팅|UV코팅)\)")
COATING_DASH_RE = re.compile(r"-\s*(양면|단면)?(무광|유광|벨벳)코팅")


def detect_coating(paper_name: str, coating_field) -> str:
    pn = paper_name or ""
    cf = (coating_field or "").strip() if coating_field else ""

    # 1) paper_name 괄호
    m = COATING_INLINE_RE.search(pn)
    if m:
        tok = m.group(1)
        if "무코팅" in tok or "비코팅" in tok: return "비코팅"
        if "무광" in tok: return "무광코팅"
        if "유광" in tok: return "유광코팅"
        if "벨벳" in tok: return "벨벳코팅"

    # 2) printcity dash
    m = COATING_DASH_RE.search(pn)
    if m:
        kind = m.group(2)
        return {"무광": "무광코팅", "유광": "유광코팅", "벨벳": "벨벳코팅"}.get(kind, "비코팅")

    # 3) coating field
    if cf:
        if "코팅없음" in cf or cf in ("무코팅", "비코팅"): return "비코팅"
        if "무광" in cf: return "무광코팅"
        if "유광" in cf: return "유광코팅"
        if "벨벳" in cf: return "벨벳코팅"
        if "홀로그램" in cf: return "홀로그램코팅"

    # 4) default
    return "비코팅"


# 평량 추출
WEIGHT_GMS_RE = re.compile(r"(\d{2,4})\s*g\s*/\s*㎡")
WEIGHT_RE = re.compile(r"(\d{2,4})\s*g(?!/)")
WEIGHT_MICRON_RE = re.compile(r"(\d{2,4})\s*[μuU](?!nique)")


def extract_weight(paper_name, paper_weight_text=None):
    """(weight_int, is_plastic). paper_name 에 없으면 paper_weight_text fallback."""
    for src in (paper_name, paper_weight_text):
        if not src: continue
        m = WEIGHT_GMS_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_MICRON_RE.search(src)
        if m: return (int(m.group(1)), True)
    return (None, False)


# ── 피드백 xlsx ──

def load_feedback():
    wb = load_workbook(FEEDBACK, data_only=True)
    ws = wb["명함 용지"]
    rows = []
    for r in range(2, ws.max_row + 1):
        canon = ws.cell(r, 1).value
        gram = ws.cell(r, 2).value
        if canon is None and gram is None: continue
        canon = (canon or "").strip()
        sites = {
            "printcity": _split_aliases(ws.cell(r, 3).value),
            "swadpia":   _split_aliases(ws.cell(r, 4).value),
            "wowpress":  _split_aliases(ws.cell(r, 5).value),
            "adsland":   _split_aliases(ws.cell(r, 6).value),
            "dtpia":     _split_aliases(ws.cell(r, 7).value),
        }
        memo = " | ".join(filter(None, [_s(ws.cell(r, 8).value), _s(ws.cell(r, 9).value)]))
        rows.append({"canonical": canon, "weight_range": _s(gram), "sites": sites, "memo": memo})
    return rows


def _s(v):
    return "" if v is None else str(v).strip()


def _split_aliases(v):
    if v is None: return []
    return [p.strip() for p in str(v).split("/") if p.strip()]


# ── 매칭 ──

def build_reverse_map(feedback):
    rev = defaultdict(dict)  # site → alias → fb_row
    for row in feedback:
        for site, aliases in row["sites"].items():
            for a in aliases:
                rev[site][a] = row
                # 평량 strip
                strip1 = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", a).strip()
                strip1 = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", strip1).strip()
                if strip1 and strip1 != a:
                    rev[site].setdefault(strip1, row)
    return rev


def normalize_for_match(site: str, paper_name: str) -> list[str]:
    cands = [paper_name]
    pn = paper_name

    # adsland: "PaperSort X Y g/㎡" → "X Yg" / "X"
    if site == "adsland":
        m = re.match(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+(.+?)\s+(\d{2,4})\s*g\s*/\s*㎡\s*$", pn)
        if m:
            cands.append(f"{m.group(2)} {m.group(3)}g")
            cands.append(m.group(2))

    # printcity dash 기반 + 코팅 strip
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?[, ].*$", "", pn).strip()
    pn2 = re.sub(r"-\s*(양면|단면)?(무광|유광|벨벳|무|비)코팅?\s*$", "", pn2).strip()
    if pn2 != pn: cands.append(pn2)

    # printcity dash → space
    pn3 = re.sub(r"-(\d{2,4}g)", r" \1", pn)
    if pn3 != pn: cands.append(pn3)

    # 코팅 괄호 제거
    pn4 = COATING_INLINE_RE.sub("", pn).replace("  ", " ").strip()
    if pn4 != pn: cands.append(pn4)

    # 평량 suffix 제거
    base = re.sub(r"\s*\d{2,4}\s*[gμu]\S*\s*$", "", pn).strip()
    base = re.sub(r"\s*\d{2,4}\s*g\s*/\s*㎡\s*$", "", base).strip()
    if base and base != pn: cands.append(base)

    seen = set(); out = []
    for c in cands:
        if c not in seen:
            seen.add(c); out.append(c)
    return out


def match_canonical(site, paper_name, weight, rev, feedback):
    """(canonical, weight_violation_bool, memo)."""
    cands = normalize_for_match(site, paper_name)
    site_map = rev.get(site, {})

    # 1) same-site exact
    for cand in cands:
        if cand in site_map:
            row = site_map[cand]
            canonical = normalize_canonical_name(row["canonical"])
            viol = check_violation(weight, row["weight_range"])
            return (canonical, viol, row["memo"])

    # 2) cross-site
    for cand in cands:
        for fb_row in feedback:
            for s, aliases in fb_row["sites"].items():
                if cand in aliases:
                    canonical = normalize_canonical_name(fb_row["canonical"])
                    viol = check_violation(weight, fb_row["weight_range"])
                    return (canonical, viol, f"cross({s}) | {fb_row['memo']}")

    return (None, False, "")


def check_violation(weight, range_str) -> bool:
    if weight is None or not range_str: return False
    rs = range_str.replace("(평량없음)", "").strip()
    if not rs: return False
    try:
        if "~" in rs:
            lo, hi = [int(float(x.strip())) for x in rs.split("~", 1)]
            return not (lo - 20 <= weight <= hi + 20)
        else:
            target = int(float(rs))
            return abs(weight - target) > 20
    except ValueError:
        return False


# ── 메인 ──

SITES = ["printcity", "swadpia", "wowpress", "adsland", "dtpia"]
SITE_LABEL = {
    "printcity": "프린트시티",
    "swadpia": "성원애드피아",
    "wowpress": "와우프레스",
    "adsland": "애즈랜드",
    "dtpia": "디티피아",
}


def aggregate_by_canonical(category, feedback, rev):
    """raw 를 (canonical, weight, coating) 기준 집계.

    반환 [{canonical, weight, coating, sites: {site: set(paper_names)}, viol, memos}]
    """
    table = defaultdict(lambda: {
        "sites": defaultdict(set),
        "viol": False,
        "memos": set(),
    })
    misses = []  # (site, paper_name, coating_norm, weight)

    for fn in sorted(os.listdir(os.path.join(ROOT, "output"))):
        if not fn.endswith("_raw_now.json"): continue
        if f"_{category}_" not in fn: continue
        site = fn.split("_")[0]
        d = json.load(open(os.path.join(ROOT, "output", fn), encoding="utf-8"))
        for i in d.get("items", []):
            pn = i.get("paper_name") or ""
            pwt = i.get("paper_weight_text") or ""
            coat_raw = i.get("coating")
            weight, is_plastic = extract_weight(pn, pwt)
            coat_norm = detect_coating(pn, coat_raw)

            if is_plastic:
                # 플라스틱 — 매칭 제외
                misses.append((site, pn, coat_norm, weight, "플라스틱"))
                continue

            canonical, viol, memo = match_canonical(site, pn, weight, rev, feedback)
            if canonical is None:
                misses.append((site, pn, coat_norm, weight, "미매칭"))
                continue

            key = (canonical, weight or 0, coat_norm)
            table[key]["sites"][site].add(pn)
            if viol: table[key]["viol"] = True
            if memo: table[key]["memos"].add(memo[:60])

    return table, misses


def write_sheet(ws, title, table, misses):
    ws.title = title
    headers = ["대표용지", "평량g", "코팅"] + [SITE_LABEL[s] for s in SITES] + ["메모"]
    ws.append(headers)
    # 정렬: canonical → weight → coating
    keys = sorted(table.keys(), key=lambda k: (k[0], k[1], k[2]))
    for k in keys:
        canonical, weight, coating = k
        info = table[k]
        row = [canonical, weight, coating]
        for s in SITES:
            papers = info["sites"].get(s, set())
            row.append(" / ".join(sorted(papers)) if papers else "")
        row.append(" | ".join(sorted(info["memos"])) if info["memos"] else "")
        ws.append(row)
        # 위반 행 노란색
        if info["viol"]:
            r = ws.max_row
            yellow = PatternFill("solid", fgColor="FFF2CC")
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = yellow

    # 미매칭 / 플라스틱 섹션
    ws.append([])
    ws.append(["── 미매칭 / 플라스틱 (canonical 신규 등록 또는 제외 결정 필요) ──"])
    ws.append(["사유", "사이트", "raw paper_name", "코팅(추정)", "평량g(추출)"])
    for site, pn, coat, w, reason in sorted(misses):
        ws.append([reason, SITE_LABEL.get(site, site), pn, coat, w if w else ""])

    # 스타일
    hf = Font(bold=True, color="FFFFFF"); hp = PatternFill("solid", fgColor="4472C4")
    ha = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hp; cell.alignment = ha
    widths = [22, 8, 12, 30, 30, 30, 30, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    feedback = load_feedback()
    rev = build_reverse_map(feedback)
    print(f"피드백 행: {len(feedback)}")

    wb = Workbook()
    for cat, sheet_name in (("card_offset", "card_offset"), ("card_digital", "card_digital")):
        table, misses = aggregate_by_canonical(cat, feedback, rev)
        if cat == "card_offset":
            ws = wb.active
        else:
            ws = wb.create_sheet()
        write_sheet(ws, sheet_name, table, misses)
        viol = sum(1 for v in table.values() if v["viol"])
        print(f"  {sheet_name}: canonical_keys={len(table)} viol={viol} miss/plastic={len(misses)}")

    wb.save(OUT_PATH)
    print(f"\n✅ saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
