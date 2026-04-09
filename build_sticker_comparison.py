"""스티커 3사 비교 엑셀 생성 (용지+사이즈 기준)"""
import json, glob, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# === 데이터 로드 ===
with open('output/ecard21_sticker_prices.json', 'r', encoding='utf-8') as f:
    ec_data = json.load(f)
ec_all = []
for pname, records in ec_data['products'].items():
    ec_all.extend(records)

bz_files = sorted(glob.glob('output/bizhows_sticker_*.json'))
with open(bz_files[-1], 'r', encoding='utf-8') as f:
    bz_data = json.load(f)
bz_all = bz_data.get('records', bz_data) if isinstance(bz_data, dict) else bz_data

oh_files = sorted(glob.glob('output/ohprint_sticker_*.json'))
with open(oh_files[-1], 'r', encoding='utf-8') as f:
    oh_all = json.load(f)


def parse_bz_price(price_str):
    if not price_str:
        return None
    prices = re.findall(r'([\d,]+)원', str(price_str))
    if len(prices) >= 2:
        return int(prices[-1].replace(',', ''))
    elif len(prices) == 1:
        return int(prices[0].replace(',', ''))
    return None


def find_bz_match(paper, size, shape_hint=''):
    paper_lower = paper.lower()
    for bz in bz_all:
        bz_paper = bz.get('paper', '')
        bz_paper_lower = bz_paper.lower()
        bz_shape = bz.get('shape', '')

        matched = False
        if '아트지' in paper_lower and '유광' in paper_lower and '초강접' not in paper_lower:
            if '코팅아트지' in bz_paper_lower or '유광아트지' in bz_paper_lower:
                matched = True
        elif '아트지' in paper_lower and '무광' in paper_lower and '초강접' not in paper_lower:
            if '무광아트지' in bz_paper_lower or '무코팅아트지' in bz_paper_lower:
                matched = True
        elif '비코팅' in paper_lower and '아트지' in paper_lower:
            if '무코팅아트지' in bz_paper_lower:
                matched = True
        elif '초강접' in paper_lower:
            if '초강접' in bz_paper_lower:
                matched = True
        elif '모조지' in paper_lower:
            if '모조지' in bz_paper_lower:
                matched = True
        elif '유포지' in paper_lower:
            if '유포지' in bz_paper_lower:
                matched = True
        elif '크라프트' in paper_lower:
            if '크라프트' in bz_paper_lower:
                matched = True
        elif '은데드롱' in paper_lower:
            if '프리미엄투명' in bz_paper_lower or '은데드롱' in bz_paper_lower:
                matched = True
        elif '투명' in paper_lower and '데드롱' in paper_lower:
            if '투명데드롱' in bz_paper_lower:
                matched = True

        if not matched:
            continue

        # 모양 매칭: 원형↔원형, 사각↔사각형, 직사각↔직사각형
        shape_ok = False
        if '원형' in size:
            shape_ok = '원형' in bz_shape
        elif '정사각' in size:
            shape_ok = '사각' in bz_shape and '직' not in bz_shape
        elif '직사각' in size or '명함' in size:
            shape_ok = '직사각' in bz_shape

        if shape_ok:
            return parse_bz_price(bz.get('price', '')), bz_paper, bz.get('size', ''), bz_shape

    return None, None, None, None


def find_oh_match(paper, size):
    paper_lower = paper.lower()
    # 오프린트미 매칭 (확인된 것만)
    target_oh = None
    if '아트지' in paper_lower and ('유광' in paper_lower or '유광코팅' in paper_lower) and '초강접' not in paper_lower:
        target_oh = '소프트'
    elif '아트지' in paper_lower and '무광' in paper_lower and '초강접' not in paper_lower:
        target_oh = '스탠다드'
    elif '모조지' in paper_lower:
        target_oh = '화이트 페이퍼'
    elif '유포지' in paper_lower:
        target_oh = '화이트 플라스틱'
    elif '투명' in paper_lower and '데드롱' in paper_lower:
        target_oh = '투명'
    elif '은데드롱' in paper_lower:
        target_oh = '홀로그램'
    # 리무버블 = 명함천국에 없음, 크라프트 = 오프린트미에 없음

    if not target_oh:
        return None, None, None

    # 사이즈에서 숫자 추출
    size_nums = re.findall(r'(\d+)', size)

    for oh in oh_all:
        oh_paper = oh.get('용지', oh.get('paper', ''))
        oh_size = oh.get('사이즈', oh.get('size', ''))
        oh_shape = oh.get('형태', oh.get('shape', ''))

        if target_oh.lower() not in oh_paper.lower():
            continue

        oh_size_nums = re.findall(r'(\d+)', oh_size)

        # 원형: 50mm ↔ 50X50
        if '원형' in size and '원형' in oh_shape:
            if size_nums and oh_size_nums and size_nums[0] == oh_size_nums[0]:
                price = oh.get('판매가', oh.get('price_1000'))
                return price, oh_paper, f"{oh_shape} {oh_size}"

        # 정사각: 50x50 ↔ 50X50
        if '정사각' in size and '정사각' in oh_shape:
            if size_nums and oh_size_nums and size_nums[0] == oh_size_nums[0]:
                price = oh.get('판매가', oh.get('price_1000'))
                return price, oh_paper, f"{oh_shape} {oh_size}"

        # 직사각/명함: 50x80 ↔ 유사 사이즈
        if ('직사각' in size or '명함' in size) and '직사각' in oh_shape:
            if len(size_nums) >= 2 and len(oh_size_nums) >= 2:
                ec_area = int(size_nums[0]) * int(size_nums[1])
                oh_area = int(oh_size_nums[0]) * int(oh_size_nums[1])
                if abs(ec_area - oh_area) / max(ec_area, oh_area) < 0.25:
                    price = oh.get('판매가', oh.get('price_1000'))
                    return price, oh_paper, f"{oh_shape} {oh_size}"

    return None, None, None


# === 엑셀 생성 ===
wb = Workbook()
ws = wb.active
ws.title = '스티커 용지+사이즈 비교'

hfw = Font(name='Arial', size=10, bold=True, color='FFFFFF')
cf = Font(name='Arial', size=10)
hf_ec = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hf_bz = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
hf_oh = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
hf_info = PatternFill(start_color='7B7B7B', end_color='7B7B7B', fill_type='solid')

ws.merge_cells('A1:J1')
ws['A1'] = '스티커 3사 가격 비교 (용지+사이즈 기준, 1,000매)'
ws['A1'].font = Font(name='Arial', size=14, bold=True)
ws.merge_cells('A2:J2')
ws['A2'] = '명함천국 제품 기준 | 비즈하우스는 65mm만 수집(60mm행 참고) | VAT: 명함천국 포함, 타사 별도'
ws['A2'].font = Font(name='Arial', size=10, color='666666')

headers = [
    ('용지', hf_info), ('사이즈', hf_info), ('제품유형', hf_info),
    ('명함천국 가격', hf_ec),
    ('비즈하우스 용지', hf_bz), ('비즈하우스 사이즈', hf_bz), ('비즈하우스 가격', hf_bz),
    ('오프린트미 용지', hf_oh), ('오프린트미 사이즈', hf_oh), ('오프린트미 가격', hf_oh),
]
for col, (h, fill) in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = hfw
    c.fill = fill
    c.alignment = Alignment(horizontal='center')

seen = set()
row = 5
fills_list = ['E8F0FE', 'FFFFFF']
ci = 0
prev_paper = None

for r in ec_all:
    key = (r['paper_code'], r['size_label'])
    if key in seen:
        continue
    seen.add(key)

    paper = r['paper_name']
    size = r['size_label']

    if paper != prev_paper:
        if prev_paper is not None:
            ci = 1 - ci
        prev_paper = paper
    fill = PatternFill(start_color=fills_list[ci], end_color=fills_list[ci], fill_type='solid')

    bz_price, bz_paper, bz_size, bz_shape = find_bz_match(paper, size)
    oh_price, oh_paper, oh_size = find_oh_match(paper, size)

    ws.cell(row=row, column=1, value=paper).font = cf
    ws.cell(row=row, column=2, value=size).font = cf
    ws.cell(row=row, column=3, value=r['product_name']).font = cf

    pc = ws.cell(row=row, column=4, value=r['price'])
    pc.font = cf
    pc.number_format = '#,##0'
    pc.alignment = Alignment(horizontal='right')

    ws.cell(row=row, column=5, value=bz_paper or '-').font = cf
    bz_size_display = f"{bz_shape} {bz_size}" if bz_shape else (bz_size or '-')
    ws.cell(row=row, column=6, value=bz_size_display).font = cf
    if bz_price:
        c = ws.cell(row=row, column=7, value=bz_price)
        c.number_format = '#,##0'
    else:
        c = ws.cell(row=row, column=7, value='-')
    c.font = cf
    c.alignment = Alignment(horizontal='right')

    ws.cell(row=row, column=8, value=oh_paper or '-').font = cf
    ws.cell(row=row, column=9, value=oh_size or '-').font = cf
    if oh_price:
        c = ws.cell(row=row, column=10, value=oh_price)
        c.number_format = '#,##0'
    else:
        c = ws.cell(row=row, column=10, value='-')
    c.font = cf
    c.alignment = Alignment(horizontal='right')

    for c2 in range(1, 11):
        ws.cell(row=row, column=c2).fill = fill
    row += 1

# === Sheet 2: 오프린트미 참고 (카테고리별 가격) ===
ws2 = wb.create_sheet('오프린트미 참고')
ws2.merge_cells('A1:F1')
ws2['A1'] = '오프린트미 스티커 가격 (카테고리별 참고용)'
ws2['A1'].font = Font(name='Arial', size=12, bold=True)
ws2.merge_cells('A2:F2')
ws2['A2'] = '오프린트미는 용지 실명 비공개 (소프트/스탠다드 등 자체 브랜드명). 정확한 용지 매칭 불가 → 참고만'
ws2['A2'].font = Font(name='Arial', size=10, color='FF0000')

oh_headers = [('카테고리', hf_oh), ('형태', hf_oh), ('사이즈', hf_oh), ('수량', hf_oh), ('가격(원)', hf_oh), ('정가(원)', hf_oh)]
for col, (h, fill_h) in enumerate(oh_headers, 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font = hfw
    c.fill = fill_h
    c.alignment = Alignment(horizontal='center')

oh_row = 5
oh_seen = set()
ci2 = 0
prev_oh_paper = None
for r in oh_all:
    oh_paper = r.get('용지', '')
    oh_shape = r.get('형태', '')
    oh_size = r.get('사이즈', '')
    key = (oh_paper, oh_shape, oh_size)
    if key in oh_seen:
        continue
    oh_seen.add(key)

    if oh_paper != prev_oh_paper:
        if prev_oh_paper is not None:
            ci2 = 1 - ci2
        prev_oh_paper = oh_paper
    fill2 = PatternFill(start_color=fills_list[ci2], end_color=fills_list[ci2], fill_type='solid')

    ws2.cell(row=oh_row, column=1, value=oh_paper).font = cf
    ws2.cell(row=oh_row, column=2, value=oh_shape).font = cf
    ws2.cell(row=oh_row, column=3, value=oh_size).font = cf
    ws2.cell(row=oh_row, column=4, value=r.get('수량', 1000)).font = cf
    ws2.cell(row=oh_row, column=4).alignment = Alignment(horizontal='center')

    sell = r.get('판매가')
    if sell:
        c = ws2.cell(row=oh_row, column=5, value=sell)
        c.number_format = '#,##0'
    else:
        c = ws2.cell(row=oh_row, column=5, value='-')
    c.font = cf
    c.alignment = Alignment(horizontal='right')

    orig = r.get('정가')
    if orig:
        c = ws2.cell(row=oh_row, column=6, value=orig)
        c.number_format = '#,##0'
    else:
        c = ws2.cell(row=oh_row, column=6, value='-')
    c.font = cf
    c.alignment = Alignment(horizontal='right')

    for c2 in range(1, 7):
        ws2.cell(row=oh_row, column=c2).fill = fill2
    oh_row += 1

wb.save('output/sticker_reference_comparison.xlsx')
print(f"엑셀 저장: Sheet1={row - 5}행, Sheet2(오프린트미 참고)={oh_row - 5}행")
