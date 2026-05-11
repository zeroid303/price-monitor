"""프린트시티 봉투 엑셀 → config/targets/envelope.yaml printcity 섹션 자동 생성.

입력:
  - data/printcity/envelope_color.xlsx (칼라봉투, 단면4도, 26 paper × 6 규격size)

수집 정책 (2026-05-11~ 결정):
  - 컬러봉투 (단면4도) 만 수집. 흑백/기성봉투(단면1도) 제외.
  - SIZ_EVST + ENP:ST (또는 ENP 없음) 만 추출 — 자켓/창봉투/도무송 제외
  - 표준 매수 = 1000매

target item 스키마:
  - product: '칼라봉투'
  - paper, paper_code
  - size, size_code, size_label
  - print_mode, color_code
  - qty (매수, 정수)
  - price (공급가)
"""
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC_COLOR = ROOT / "data/printcity/envelope_color.xlsx"
DST = ROOT / "config/targets/envelope.yaml"

# 표준 매수 (다른 사이트 raw 도 모두 1000매)
TARGET_QTY_MAE = [1000]

# size canonical 매핑 (EVST 만)
SIZE_CODE_TO_LABEL = {
    "SIZ_EVST:330x245": ("대봉투", "330x245"),
    "SIZ_EVST:255x190": ("9절봉투", "255x190"),
    "SIZ_EVST:220x105": ("소봉투", "220x105"),
    "SIZ_EVST:305x215": ("6절봉투", "305x215"),
    "SIZ_EVST:290x390": ("8절봉투", "290x390"),
    "SIZ_EVST:330x450": ("A3봉투", "330x450"),
}


def parse_xlsx(path: Path, default_product: str, default_color: str):
    """엑셀 → list[dict] 추출. SIZ_EVST + (ENP=ST 또는 없음) 만.

    각 dict: product, paper, paper_code, size, size_label, size_code,
             print_mode, color_code, qty, price
    """
    if not path.exists():
        print(f"  ⚠ {path} 없음 — skip")
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []
    cur_paper = cur_paper_code = None
    cur_size_code = None
    cur_color = cur_color_code = None
    cur_ent = None    # 봉투타입 (ST/JK)
    cur_enp = None    # 가공 (ST/DM)

    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        qty = ws.cell(r, 3).value
        value = ws.cell(r, 5).value

        if isinstance(code, str) and ":" in code:
            if code.startswith("ENT:"):
                cur_ent = code.split(":", 1)[1]
                continue
            if code.startswith("ENP:"):
                cur_enp = code.split(":", 1)[1]
                continue
            if code.startswith("MAT:"):
                cur_paper, cur_paper_code = title, code
                continue
            if code.startswith("SIZ_"):
                cur_size_code = code
                continue
            if code.startswith("COL:"):
                cur_color, cur_color_code = title, code
                continue

        # data row (qty + value)
        if not (isinstance(qty, (int, float)) and isinstance(value, (int, float))):
            continue
        qty_mae = int(qty)
        if qty_mae not in TARGET_QTY_MAE:
            continue
        # 필터 — 규격형 + 일반가공만 (자켓/창봉투/도무송 제외)
        if cur_size_code not in SIZE_CODE_TO_LABEL:
            continue
        if cur_ent is not None and cur_ent != "ST":
            continue  # 자켓형 제외
        if cur_enp is not None and cur_enp != "ST":
            continue  # 도무송 제외

        size_canonical, size_label = SIZE_CODE_TO_LABEL[cur_size_code]
        items.append({
            "product": default_product,
            "paper": cur_paper,
            "paper_code": cur_paper_code,
            "size": size_canonical,
            "size_label": size_label,
            "size_code": cur_size_code,
            "print_mode": cur_color or default_color,
            "color_code": cur_color_code,
            "qty": qty_mae,
            "price": int(value) if value else None,
        })
    return items


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    items = parse_xlsx(SRC_COLOR, default_product="칼라봉투", default_color="단면4도")

    # 기존 yaml read or new
    if DST.exists():
        data = yaml.safe_load(DST.read_text(encoding="utf-8")) or {}
    else:
        data = {}

    data["printcity"] = {
        "_description": "정적 엑셀 source — data/printcity/envelope_color.xlsx (단면4도 칼라봉투 전용)",
        "sources": ["data/printcity/envelope_color.xlsx"],
        "filters_applied": {
            "qty_mae": TARGET_QTY_MAE,
            "size_codes": list(SIZE_CODE_TO_LABEL.keys()),
            "envelope_types": ["ST"],
            "processing": ["ST", "(none)"],
            "color_codes": ["COL:40"],
        },
        "price_vat_included": False,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "items": items,
    }

    DST.parent.mkdir(parents=True, exist_ok=True)
    DST.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False,
                       default_flow_style=False, width=140),
        encoding="utf-8",
    )

    from collections import Counter
    by_product = Counter(it["product"] for it in items)
    by_size = Counter(it["size"] for it in items)
    null_count = sum(1 for it in items if it["price"] is None)
    print(f"✅ {DST}: printcity 섹션 갱신")
    print(f"  총 {len(items)} items")
    print(f"  product 분포: {dict(by_product)}")
    print(f"  size 분포: {dict(by_size)}")
    print(f"  미공급(price=null): {null_count}건")


if __name__ == "__main__":
    main()
