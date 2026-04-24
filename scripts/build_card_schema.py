"""피드백 엑셀 → config/schemas/card.json 자동 변환.

입력: 다운로드 폴더의 '경쟁사 가격 모니터링 피드백.xlsx' (명함 용지 시트)
    컬럼: 대표용지 | 평량(g) | 프린트시티 | 성원애드피아 | 와우프레스 | 디티피아 | 애즈랜드 | 디티피아 | 비고 | 비고
    두 번째 '디티피아' 컬럼(col7)은 col5와 실질 동일한 중복 — 무시.

출력: config/schemas/card.json
    - _normalization.paper_name.canonical: {이름: {weights[], weight_ranges[], weightless, aliases_by_site{}, notes[]}}
    - _normalization.paper_name.aliases: 기존 normalize.py 호환용 flat {canonical: [base_aliases]}
    - 나머지 섹션(coating/print_mode/size/qty/paper_name_extract) 은 기존 card_mapping_rule.json 에서 그대로 이관.
    - _match_axes.axes: 6축 (paper_name + paper_weight_g + coating + print_mode + size + qty)
    - sites 섹션은 제거 (config/sites/{site}.yaml 로 이전 예정).

사용:
    python -m scripts.build_card_schema
    python -m scripts.build_card_schema --input "경로/to/피드백.xlsx" --output config/schemas/card.json
"""
import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

DEFAULT_INPUT = ROOT / "data" / "reference" / "card_paper_aliases.xlsx"
DEFAULT_OUTPUT = ROOT / "config" / "schemas" / "card.yaml"
LEGACY_RULE = ROOT / "config" / "card_mapping_rule.json"

# 피드백 시트 컬럼 인덱스 → 사이트 키
SITE_COLS: list[tuple[int, str]] = [
    (2, "printcity"),
    (3, "swadpia"),
    (4, "wowpress"),
    (5, "dtpia"),
    (6, "adsland"),
    # col7 은 col5(디티피아)와 실질 중복 — 무시
]

NOISE_WEIGHT_RE = re.compile(r"(\d+)\s*g.*$", re.IGNORECASE)
# 알리아스 구분자: '/' (공백 유무 상관없음) 와 줄바꿈 모두 허용
SPLIT_RE = re.compile(r"\s*[/\n]\s*")


def parse_weight_cell(cell: Any) -> dict:
    """평량 셀 해석. 단일/범위/복수/(평량없음) 모두 처리."""
    if cell is None:
        return {"weights": [], "ranges": [], "weightless": False}
    if isinstance(cell, (int, float)):
        return {"weights": [int(cell)], "ranges": [], "weightless": False}
    s = str(cell).strip()
    if s == "" or s == "-":
        return {"weights": [], "ranges": [], "weightless": False}
    if "평량없음" in s:
        return {"weights": [], "ranges": [], "weightless": True}
    weights: list[int] = []
    ranges: list[list[int]] = []
    for part in re.split(r"[,/]", s):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(\d+)\s*[~\-]\s*(\d+)$", part)
        if m:
            ranges.append([int(m.group(1)), int(m.group(2))])
            continue
        m = re.match(r"^(\d+)$", part)
        if m:
            weights.append(int(m.group(1)))
    return {
        "weights": sorted(set(weights)),
        "ranges": ranges,
        "weightless": False,
    }


def parse_aliases(cell: Any) -> list[str]:
    if cell is None:
        return []
    if not isinstance(cell, str):
        cell = str(cell)
    parts = [p.strip() for p in SPLIT_RE.split(cell) if p and p.strip()]
    return parts


def extract_base(alias: str) -> str:
    """alias에서 trailing '숫자g' 노이즈를 떼어 base 이름만. normalize.py 호환용."""
    s = NOISE_WEIGHT_RE.sub("", alias).strip()
    # 괄호 속 코팅 표기 제거 — normalize 단계에서 다시 처리하므로 base alias에는 없어야.
    s = re.sub(r"\((무광코팅|유광코팅|벨벳코팅|무코팅)\)", "", s).strip()
    # 연속 공백 정리
    s = re.sub(r"\s+", " ", s).strip()
    # 끝 구분자 제거
    s = s.rstrip("-_ ").strip()
    return s


def load_feedback(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    sheet_name = "명함 용지"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")
    ws = wb[sheet_name]

    canonical: dict[str, dict] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name:
            continue

        winfo = parse_weight_cell(row[1] if len(row) > 1 else None)
        note_cell = row[8] if len(row) > 8 else None

        entry = canonical.setdefault(name, {
            "weights": set(),
            "weight_ranges": [],
            "weightless": False,
            "aliases_by_site": {},
            "notes": set(),
        })
        entry["weights"].update(winfo["weights"])
        for r in winfo["ranges"]:
            if r not in entry["weight_ranges"]:
                entry["weight_ranges"].append(r)
        entry["weightless"] = entry["weightless"] or winfo["weightless"]

        for col_idx, site_key in SITE_COLS:
            if col_idx >= len(row):
                continue
            for alias in parse_aliases(row[col_idx]):
                entry["aliases_by_site"].setdefault(site_key, set()).add(alias)

        if note_cell and isinstance(note_cell, str) and note_cell.strip():
            entry["notes"].add(note_cell.strip())

    wb.close()

    # 직렬화 가능한 형태로 변환
    out = {}
    for name in sorted(canonical):
        info = canonical[name]
        out[name] = {
            "weights": sorted(info["weights"]),
            "weight_ranges": info["weight_ranges"],
            "weightless": info["weightless"],
            "aliases": {
                site: sorted(list(aliases))
                for site, aliases in sorted(info["aliases_by_site"].items())
            },
        }
        if info["notes"]:
            out[name]["notes"] = sorted(list(info["notes"]))
    return out


def build_flat_aliases(canonical_map: dict) -> dict[str, list[str]]:
    """기존 normalize.py가 요구하는 {canonical: [base_aliases]} 형태 생성.

    각 사이트 alias에서 trailing 평량/코팅 표기를 떼고 dedupe.
    """
    flat: dict[str, list[str]] = {}
    for canon, info in canonical_map.items():
        bases: set[str] = set()
        # canonical 이름 자체도 포함 (normalize._build_alias_lookup 이 canonical→canonical 자기참조 필요)
        bases.add(canon)
        for aliases in info["aliases"].values():
            for a in aliases:
                b = extract_base(a)
                if b and b != "":
                    bases.add(b)
        # canonical과 동일한 것도 유지 (중복 허용)
        flat[canon] = sorted(bases)
    return flat


def build_schema(canonical_map: dict, legacy: dict) -> dict:
    """최종 schemas/card.json 조립."""
    flat_aliases = build_flat_aliases(canonical_map)

    # legacy 의 _normalization 에서 paper_name 만 교체, 나머지 그대로
    legacy_norm = legacy.get("_normalization", {})
    new_paper_name_rule = {
        "_description": (
            "용지명 정규화. noise strip → coating 추출 → canonical alias 적용 → '{canonical} {weight}g' 형태. "
            "canonical 구조는 피드백 엑셀(경쟁사 가격 모니터링 피드백.xlsx) 기반으로 자동 생성. "
            "aliases(flat)은 기존 normalize.py 호환용 역 인덱스."
        ),
        "canonical": canonical_map,
        "aliases": flat_aliases,
        "noise_suffix_regex": legacy_norm.get("paper_name", {}).get(
            "noise_suffix_regex", r"(\d+)\s*g.*$"
        ),
    }

    schema: dict = {
        "_description": (
            "명함 카테고리 정규화 규칙 (canonical 사전 + alias). "
            "사이트 메타데이터(name/base_url/vat_included 등)는 config/sites/{site}.yaml 참조."
        ),
        "_match_axes": {
            "_description": (
                "매칭 축 6개. paper_name(canonical 이름)과 paper_weight_g(평량 수치)를 분리해 "
                "±N g 허용 같은 산술 매칭을 용이하게 함."
            ),
            "axes": [
                "paper_name",
                "paper_weight_g",
                "coating",
                "print_mode",
                "size",
                "qty",
            ],
            "weight_tolerance_g": 25,
        },
        "_defaults": legacy.get("_defaults", {"match_method": "substring", "qty_default": 200}),
        "_normalization": {
            "paper_name": new_paper_name_rule,
            "coating": legacy_norm.get("coating", {}),
            "print_mode": legacy_norm.get("print_mode", {}),
            "size": legacy_norm.get("size", {}),
            "qty": legacy_norm.get("qty", {}),
            "paper_name_extract": legacy_norm.get("paper_name_extract", {}),
        },
    }
    return schema


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT,
                    help="피드백 엑셀 경로")
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                    help="schemas/card.yaml 출력 경로")
    ap.add_argument("--legacy", type=Path, default=LEGACY_RULE,
                    help="기존 card_mapping_rule.json (coating/print_mode/size/qty 이관)")
    args = ap.parse_args()

    if not args.input.exists():
        raise SystemExit(f"입력 파일 없음: {args.input}")
    if not args.legacy.exists():
        raise SystemExit(f"legacy 파일 없음: {args.legacy}")

    legacy = json.loads(args.legacy.read_text(encoding="utf-8"))
    canonical_map = load_feedback(args.input)
    schema = build_schema(canonical_map, legacy)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    # dict 삽입 순서 그대로 덤프 (sort_keys=False), 한글 그대로, 블록 스타일
    args.output.write_text(
        yaml.safe_dump(
            schema,
            allow_unicode=True,
            sort_keys=False,
            default_flow_style=False,
            width=120,
        ),
        encoding="utf-8",
    )

    # 리포트
    total_aliases = sum(
        sum(len(a) for a in info["aliases"].values())
        for info in canonical_map.values()
    )
    site_coverage: dict[str, int] = {}
    for info in canonical_map.values():
        for site, aliases in info["aliases"].items():
            if aliases:
                site_coverage[site] = site_coverage.get(site, 0) + 1

    print(f"✅ schemas 생성: {args.output}")
    print(f"  canonical: {len(canonical_map)} 용지")
    print(f"  total aliases: {total_aliases}")
    print(f"  site coverage (canonical 중 해당 사이트 alias 있는 수):")
    for site, cnt in sorted(site_coverage.items()):
        print(f"    {site}: {cnt}")
    weightless = [n for n, info in canonical_map.items() if info["weightless"]]
    if weightless:
        print(f"  weightless 용지 ({len(weightless)}): {weightless}")


if __name__ == "__main__":
    main()
