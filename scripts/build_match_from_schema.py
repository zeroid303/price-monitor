"""새 schema (config/schemas/card_offset.yaml + card_digital.yaml) 기반 매칭표 생성.

각 행 = (대표용지, 평량g, 코팅) canonical key.
컬럼 = 사이트별 raw paper_name list.
"""
import json
import os
import re
import sys
from collections import defaultdict

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.expanduser("~/Downloads/명함_용지_매칭표_2026-04-30.xlsx")

SITES = ["printcity", "swadpia", "wowpress", "adsland", "dtpia"]
SITE_LABEL = {
    "printcity": "프린트시티",
    "swadpia":   "성원애드피아",
    "wowpress":  "와우프레스",
    "adsland":   "애즈랜드",
    "dtpia":     "디티피아",
}

WEIGHT_GMS_RE = re.compile(r"(\d{2,4})\s*g\s*/\s*㎡")
WEIGHT_RE = re.compile(r"(\d{2,4})\s*g(?!/)")
WEIGHT_MICRON_RE = re.compile(r"(\d{2,4})\s*[μuU](?!nique)")

COATING_INLINE_RE = re.compile(r"\((무광코팅|유광코팅|벨벳코팅|무코팅|비코팅|UV유광코팅|UV코팅)\)")
COATING_DASH_RE = re.compile(r"-\s*(양면|단면)?(무광|유광|벨벳)코팅")


def extract_weight(paper_name, paper_weight_text=None):
    for src in (paper_name, paper_weight_text):
        if not src: continue
        m = WEIGHT_GMS_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_RE.search(src)
        if m: return (int(m.group(1)), False)
        m = WEIGHT_MICRON_RE.search(src)
        if m: return (int(m.group(1)), True)
    return (None, False)


def detect_coating(paper_name, coating_field):
    pn = paper_name or ""
    cf = (coating_field or "").strip() if coating_field else ""
    m = COATING_INLINE_RE.search(pn)
    if m:
        tok = m.group(1)
        if "무코팅" in tok or "비코팅" in tok: return "비코팅"
        if "무광" in tok: return "무광코팅"
        if "유광" in tok: return "유광코팅"
        if "벨벳" in tok: return "벨벳코팅"
    m = COATING_DASH_RE.search(pn)
    if m:
        return {"무광": "무광코팅", "유광": "유광코팅", "벨벳": "벨벳코팅"}.get(m.group(2), "비코팅")
    if cf:
        if "코팅없음" in cf or cf in ("무코팅", "비코팅"): return "비코팅"
        if "무광" in cf: return "무광코팅"
        if "유광" in cf: return "유광코팅"
        if "벨벳" in cf: return "벨벳코팅"
        if "홀로그램" in cf: return "홀로그램코팅"
    return "비코팅"


def load_schema(category):
    """schema yaml → {alias: (canonical, weight_g)} 역인덱스."""
    fp = os.path.join(ROOT, f"config/schemas/{category}.yaml")
    d = yaml.safe_load(open(fp, encoding="utf-8"))
    canon = d["_normalization"]["paper_name"]["canonical"]
    rev = defaultdict(dict)  # site → alias → (canonical, weight)
    canonical_keys = []  # all (canonical, weight) keys with site dict
    for canonical_name, info in canon.items():
        for w_key, w_info in info["weights"].items():
            weight = w_key if isinstance(w_key, int) else 0
            sites_aliases = w_info.get("aliases", {})
            canonical_keys.append({
                "canonical": canonical_name,
                "weight": weight,
                "sites": sites_aliases,
                "memo": w_info.get("memo", ""),
            })
            for site, aliases in sites_aliases.items():
                for a in aliases:
                    rev[site][a] = (canonical_name, weight)
    return canonical_keys, rev


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    wb = Workbook()
    for idx, (cat, label) in enumerate((("card_offset", "오프셋"), ("card_digital", "디지털"))):
        canonical_keys, rev = load_schema(cat)

        # raw record → 매칭. 결과: (canonical, weight, coating) → {site: set(papers)}
        table = defaultdict(lambda: {"sites": defaultdict(set), "memos": set()})
        unmatched = []

        for fn in sorted(os.listdir(os.path.join(ROOT, "output"))):
            if not fn.endswith("_raw_now.json"): continue
            if f"_{cat}_" not in fn: continue
            site = fn.split("_")[0]
            d = json.load(open(os.path.join(ROOT, "output", fn), encoding="utf-8"))
            for i in d.get("items", []):
                pn = i.get("paper_name") or ""
                pwt = i.get("paper_weight_text") or ""
                coat_raw = i.get("coating")
                weight, is_plastic = extract_weight(pn, pwt)
                coat_norm = detect_coating(pn, coat_raw)

                # 매칭 — raw paper_name 변형들 시도
                cands = [pn]
                # adsland paperSort strip
                m = re.match(r"^(일반지|고급지|펄지|친환경 재생지|색지|한지)\s+(.+?)\s+(\d{2,4})\s*g\s*/\s*㎡\s*$", pn)
                if m:
                    cands.append(f"{m.group(2)} {m.group(3)}g")
                # g/㎡ → g
                cand_g = re.sub(r"(\d+)\s*g\s*/\s*㎡", r"\1g", pn)
                if cand_g != pn:
                    cands.append(cand_g)

                hit = None
                site_map = rev.get(site, {})
                for c in cands:
                    if c in site_map:
                        hit = site_map[c]; break
                if not hit:
                    # cross-site 매칭
                    for c in cands:
                        for s2, m_dict in rev.items():
                            if c in m_dict:
                                hit = m_dict[c]; break
                        if hit: break

                if hit:
                    canonical, w_canon = hit
                    table[(canonical, w_canon, coat_norm)]["sites"][site].add(pn)
                else:
                    unmatched.append((site, pn, weight, coat_norm))

        # canonical_keys 의 entry 들 중 raw 에 없는 것도 일단 row 표시
        # (schema 의 모든 (canonical, weight) 행 표시)
        all_keys = set(table.keys())
        for ck in canonical_keys:
            for coat in ("비코팅", "무광코팅", "유광코팅", "벨벳코팅", "홀로그램코팅"):
                key = (ck["canonical"], ck["weight"], coat)
                # raw 매칭 있을 때만 행 추가 (없는 코팅은 표시 X)
                if key in table:
                    for s, aliases in ck["sites"].items():
                        # schema 의 alias 도 표시 (raw 와 다를 수 있음)
                        pass

        # 정렬: 대표용지 → 평량 → 코팅
        keys = sorted(table.keys(), key=lambda k: (k[0], k[1], k[2]))

        if idx == 0:
            ws = wb.active; ws.title = cat
        else:
            ws = wb.create_sheet(cat)
        headers = ["대표용지", "평량g", "코팅"] + [SITE_LABEL[s] for s in SITES] + ["메모"]
        ws.append(headers)
        for k in keys:
            canonical, weight, coating = k
            info = table[k]
            row = [canonical, weight, coating]
            for s in SITES:
                papers = sorted(info["sites"].get(s, set()))
                row.append(" / ".join(papers) if papers else "")
            row.append(" | ".join(sorted(info["memos"])) if info["memos"] else "")
            ws.append(row)

        # 미매칭 섹션
        ws.append([])
        ws.append([f"── 미매칭 raw paper_name (schema 에 없음, {len(unmatched)} 건) ──"])
        ws.append(["사이트", "raw paper_name", "평량g", "코팅"])
        seen = set()
        for site, pn, w, coat in sorted(unmatched):
            sig = (site, pn)
            if sig in seen: continue
            seen.add(sig)
            ws.append([SITE_LABEL[site], pn, w if w else "", coat])

        # 스타일
        hf = Font(bold=True, color="FFFFFF"); hp = PatternFill("solid", fgColor="4472C4")
        ha = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = hf; cell.fill = hp; cell.alignment = ha
        widths = [22, 8, 12, 28, 28, 28, 28, 28, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        print(f"  {cat}: canonical_keys={len(keys)} unmatched_raw={len(set((s,p) for s,p,_,_ in unmatched))}")

    wb.save(OUT_PATH)
    print(f"\n✅ saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
