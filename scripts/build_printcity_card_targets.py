"""프린트시티 엑셀 → config/targets/card_offset.yaml · card_digital.yaml 자동 생성.

입력:
  - data/printcity/card.xlsx          (일반명함, 오프셋)
  - data/printcity/premium_card.xlsx  (고급명함, 오프셋)
  - data/printcity/digital_card.xlsx  (디지털명함)

출력:
  - config/targets/card_offset.yaml   의 `printcity` 섹션  (일반+고급명함)
  - config/targets/card_digital.yaml  의 `printcity` 섹션  (디지털명함)

필터: size=SIZ:NC-90X50, qty∈{100,200,500,1000}.
value=0/None 도 포함하되 price: null (엑셀이 잘못 비어있을 수도 있음 → 크로스체크 별도).

target item 스키마:
  - product       : "일반명함" | "고급명함" | "디지털명함"
  - paper         : 엑셀 원본 이름  (ex. "스노우화이트-250g")
  - paper_code    : 엑셀 MAT 코드  (ex. "MAT:SNW-250")
  - coating       : 엑셀 원본 이름  (ex. "코팅없음")  — 디지털은 null
  - coating_code  : 엑셀 COT 코드  — 디지털은 null
  - color_mode    : 엑셀 원본 이름  (ex. "단면4도")
  - color_code    : 엑셀 COL 코드  (ex. "COL:40")
  - size          : "90x50"
  - qty           : 정수
  - price         : 공급가(VAT 제외) 정수 또는 null

사용:
  python -m scripts.build_printcity_card_targets
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterator

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

DATA_DIR = ROOT / "data" / "printcity"
TARGETS_DIR = ROOT / "config" / "targets"

TARGET_SIZE_CODE = "SIZ:NC-90X50"
TARGET_SIZE_NAME = "90x50"
TARGET_QTYS = [100, 200, 500, 1000]

# 엑셀 파일 → (카테고리 파일 이름, product 표기)
FILES = [
    ("card.xlsx",         "card_offset",  "일반명함"),
    ("premium_card.xlsx", "card_offset",  "고급명함"),
    ("digital_card.xlsx", "card_digital", "디지털명함"),
]


def parse_xlsx(xlsx_path: Path, product: str) -> Iterator[dict]:
    """엑셀 하나를 블록 단위로 파싱. 필터 통과한 가격 행을 dict로 yield.

    블록 구조:
      (옵션 라벨 행들: COT/SIZ/MAT/COL) → (가격 행들: qty, value) → 다음 블록 헤더(code='code')
    옵션 라벨이 나오면 해당 축 상태만 갱신, 가격 행이 나오면 현재 상태 + qty + value → 하나의 item.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    state = {
        "cot_code": None, "cot_name": None,
        "siz_code": None, "siz_name": None,
        "mat_code": None, "mat_name": None,
        "col_code": None, "col_name": None,
    }

    for row in ws.iter_rows(values_only=True):
        title, code, qty, _sales, value, _calc = row
        # 1) 블록 경계 헤더 ('code' 리터럴)
        if code == "code":
            continue
        # 2) 옵션 라벨 행
        if code and isinstance(code, str):
            if code.startswith("COT:"):
                state["cot_code"], state["cot_name"] = code, title
            elif code.startswith("SIZ:"):
                state["siz_code"], state["siz_name"] = code, title
            elif code.startswith("MAT:"):
                state["mat_code"], state["mat_name"] = code, title
            elif code.startswith("COL:"):
                state["col_code"], state["col_name"] = code, title
            continue
        # 3) 가격 행 (code=None, qty 존재)
        if qty is None:
            continue

        # 필터: 사이즈·수량
        if state["siz_code"] != TARGET_SIZE_CODE:
            continue
        if qty not in TARGET_QTYS:
            continue

        # value 처리: 0/None → null
        price = None
        if isinstance(value, (int, float)) and value not in (0, 0.0):
            price = int(value)

        yield {
            "product": product,
            "paper": state["mat_name"],
            "paper_code": state["mat_code"],
            "coating": state["cot_name"],       # 디지털은 None
            "coating_code": state["cot_code"],  # 디지털은 None
            "color_mode": state["col_name"],
            "color_code": state["col_code"],
            "size": TARGET_SIZE_NAME,
            "qty": int(qty),
            "price": price,
        }
    wb.close()


def build() -> dict[str, dict]:
    """카테고리별 printcity 섹션 생성."""
    by_cat: dict[str, dict] = {
        "card_offset":  {"sources": [], "products": [], "items": []},
        "card_digital": {"sources": [], "products": [], "items": []},
    }

    for fname, cat, product in FILES:
        fp = DATA_DIR / fname
        if not fp.exists():
            raise SystemExit(f"입력 파일 없음: {fp}")
        by_cat[cat]["sources"].append(f"data/printcity/{fname}")
        if product not in by_cat[cat]["products"]:
            by_cat[cat]["products"].append(product)
        items = list(parse_xlsx(fp, product))
        by_cat[cat]["items"].extend(items)

    # 최상위 메타 래핑
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    out = {}
    for cat, payload in by_cat.items():
        out[cat] = {
            "printcity": {
                "_description": (
                    f"프린트시티 {', '.join(payload['products'])} 가격 테이블. "
                    f"data/printcity/ 엑셀에서 일회성 생성 (scripts/build_printcity_card_targets.py). "
                    f"엑셀 갱신 시 스크립트 재실행."
                ),
                "sources": payload["sources"],
                "filters_applied": {
                    "size": TARGET_SIZE_CODE,
                    "qty": TARGET_QTYS,
                },
                "price_vat_included": False,
                "generated_at": generated_at,
                "items": payload["items"],
            }
        }
    return out


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--output-dir", type=Path, default=TARGETS_DIR)
    args = ap.parse_args()

    out = build()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    for cat, payload in out.items():
        target_path = args.output_dir / f"{cat}.yaml"
        target_path.write_text(
            yaml.safe_dump(
                payload,
                allow_unicode=True,
                sort_keys=False,
                default_flow_style=False,
                width=120,
            ),
            encoding="utf-8",
        )
        pc = payload["printcity"]
        n_items = len(pc["items"])
        n_null = sum(1 for it in pc["items"] if it["price"] is None)
        print(f"✅ {target_path.name}: {n_items} items (price=null: {n_null}) — products: {', '.join(pc['_description'].split()[1:3])}")


if __name__ == "__main__":
    main()
